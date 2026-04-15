"""
Macro Pulse — refresh_macro.py
Run this script to fetch all macro indicators, generate signals,
write macro_dashboard.html and macro_pulse.xlsx.
"""

import json
import os
import ssl
import sys
import warnings
from datetime import datetime, timedelta, timezone

warnings.filterwarnings("ignore")

# Fix macOS SSL certificate verification issue
# (Python installed via python.org doesn't use the system keychain by default)
try:
    import certifi
    os.environ.setdefault("SSL_CERT_FILE", certifi.where())
    os.environ.setdefault("REQUESTS_CA_BUNDLE", certifi.where())
    ssl._create_default_https_context = ssl.create_default_context
except Exception:
    pass

# ── third-party ───────────────────────────────────────────────────────────────
try:
    from fredapi import Fred
except ImportError:
    print("ERROR: fredapi not installed. Run: pip install -r requirements.txt")
    sys.exit(1)

try:
    import yfinance as yf
except ImportError:
    print("ERROR: yfinance not installed. Run: pip install -r requirements.txt")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install -r requirements.txt")
    sys.exit(1)

import config

# ─────────────────────────────────────────────────────────────────────────────
# 1. HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def load_json(path):
    if os.path.exists(path):
        with open(path, "r") as f:
            return json.load(f)
    return {}


def save_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)


def fmt(val, decimals=2, prefix="", suffix=""):
    if val is None:
        return "N/A"
    try:
        return f"{prefix}{val:,.{decimals}f}{suffix}"
    except Exception:
        return str(val)


def direction_arrow(current, previous, threshold=0.0):
    """Return ▲ / ▼ / ▬ based on change vs threshold."""
    if current is None or previous is None:
        return "▬"
    diff = current - previous
    if abs(diff) <= threshold:
        return "▬"
    return "▲" if diff > 0 else "▼"


def safe_last(series):
    """Return the last non-null value from a pandas Series, or None."""
    if series is None or len(series) == 0:
        return None
    s = series.dropna()
    return float(s.iloc[-1]) if len(s) else None


def safe_prev(series, n=1):
    """Return the nth-from-last non-null value, or None."""
    if series is None or len(series) == 0:
        return None
    s = series.dropna()
    if len(s) <= n:
        return float(s.iloc[0]) if len(s) else None
    return float(s.iloc[-(n + 1)])


def _series_to_history(series, days=365):
    """Extract last N days of data as [[date_str, value], ...] list."""
    if series is None or len(series) == 0:
        return []
    s = series.dropna()
    if len(s) == 0:
        return []
    cutoff = s.index[-1] - timedelta(days=days)
    s = s[s.index >= cutoff]
    return [[d.strftime("%Y-%m-%d"), round(float(v), 4)] for d, v in s.items()]


def _last_date(series):
    """Return the date string of the last non-null value."""
    if series is None or len(series) == 0:
        return None
    s = series.dropna()
    if len(s) == 0:
        return None
    return s.index[-1].strftime("%Y-%m-%d")


def _prev_date(series, n=1):
    """Return the date string of the nth-from-last non-null value."""
    if series is None or len(series) == 0:
        return None
    s = series.dropna()
    if len(s) <= n:
        return s.index[0].strftime("%Y-%m-%d") if len(s) > 0 else None
    return s.index[-(n + 1)].strftime("%Y-%m-%d")


# ─────────────────────────────────────────────────────────────────────────────
# 2. FRED DATA FETCH
# ─────────────────────────────────────────────────────────────────────────────

def fetch_fred_data(previous_data: dict) -> dict:
    """Fetch FRED series; return dict of indicator dicts."""
    fred = Fred(api_key=config.FRED_API_KEY)
    results = {}
    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=config.FRED_LOOKBACK_YEARS * 365)

    def _get(series_id, start=start_date):
        try:
            s = fred.get_series(series_id, observation_start=start.strftime("%Y-%m-%d"))
            return s
        except Exception as e:
            print(f"  ⚠  FRED {series_id}: {e}")
            return None

    # Fed Funds Rate ──────────────────────────────────────────────────────────
    upper = _get("DFEDTARU")
    lower = _get("DFEDTARL")
    upper_val = safe_last(upper)
    lower_val = safe_last(lower)
    upper_prev = safe_prev(upper)
    lower_prev = safe_prev(lower)
    ff_current = upper_val
    ff_prev = upper_prev
    ff_dir = direction_arrow(ff_current, ff_prev, threshold=0.01)
    results["FED_FUNDS"] = {
        "label": "Fed Funds Rate",
        "value": ff_current,
        "previous": ff_prev,
        "value_str": f"{fmt(lower_val)}–{fmt(upper_val)}%",
        "direction": ff_dir,
        "unit": "%",
        "source": "FRED",
        "category": "Rates & Monetary Policy",
        "value_date": _last_date(upper),
        "previous_date": _prev_date(upper),
        "history": _series_to_history(upper),
    }

    # Treasury Yields ─────────────────────────────────────────────────────────
    for key, sid, label in [
        ("US_2Y",  "DGS2",  "US 2Y Treasury"),
        ("US_10Y", "DGS10", "US 10Y Treasury"),
        ("US_30Y", "DGS30", "US 30Y Treasury"),
    ]:
        s = _get(sid)
        cur = safe_last(s)
        prev = safe_prev(s)
        results[key] = {
            "label": label,
            "value": cur,
            "previous": prev,
            "direction": direction_arrow(cur, prev, threshold=0.05),
            "unit": "%",
            "source": "FRED",
            "category": "Rates & Monetary Policy",
            "value_date": _last_date(s),
            "previous_date": _prev_date(s),
            "history": _series_to_history(s),
        }

    # 2s10s Spread ────────────────────────────────────────────────────────────
    s = _get("T10Y2Y")
    cur = safe_last(s)
    prev = safe_prev(s)
    results["SPREAD_2S10S"] = {
        "label": "2s10s Spread",
        "value": cur,
        "previous": prev,
        "direction": direction_arrow(cur, prev, threshold=0.02),
        "unit": "%",
        "source": "FRED",
        "category": "Rates & Monetary Policy",
        "value_date": _last_date(s),
        "previous_date": _prev_date(s),
        "history": _series_to_history(s),
    }

    # CPI YoY ─────────────────────────────────────────────────────────────────
    # Need at least 13 monthly observations for YoY.  Pull 2+ years to be safe.
    for key, sid, label in [
        ("CPI_YOY",      "CPIAUCSL", "CPI YoY"),
        ("CORE_CPI_YOY", "CPILFESL", "Core CPI YoY"),
        ("CORE_PCE_YOY", "PCEPILFE", "Core PCE YoY"),
    ]:
        s = _get(sid, start=end_date - timedelta(days=800))
        if s is not None and len(s.dropna()) >= 13:
            sd = s.dropna()
            cur = float(sd.iloc[-1])
            prev12 = float(sd.iloc[-13]) if len(sd) >= 13 else None
            prev_month = float(sd.iloc[-2]) if len(sd) >= 2 else None
            prev12_prev = float(sd.iloc[-14]) if len(sd) >= 14 else None
            yoy = (cur / prev12 - 1) * 100 if prev12 else None
            yoy_prev = ((prev_month / prev12_prev - 1) * 100
                        if prev_month and prev12_prev else None)
            # Compute YoY history for chart
            yoy_history = []
            if len(sd) >= 13:
                yoy_series = sd.pct_change(periods=12) * 100
                yoy_clean = yoy_series.dropna()
                cutoff = yoy_clean.index[-1] - timedelta(days=365) if len(yoy_clean) > 0 else None
                if cutoff is not None:
                    yoy_clean = yoy_clean[yoy_clean.index >= cutoff]
                yoy_history = [[d.strftime("%Y-%m-%d"), round(float(v), 2)]
                               for d, v in yoy_clean.items()]
            results[key] = {
                "label": label,
                "value": yoy,
                "previous": yoy_prev,
                "direction": direction_arrow(yoy, yoy_prev, threshold=0.05),
                "unit": "% YoY",
                "source": "FRED",
                "category": "Prices & Inflation",
                "value_date": _last_date(sd),
                "previous_date": _prev_date(sd),
                "history": yoy_history,
            }
        else:
            prev_val = (previous_data.get(key, {}).get("value")
                        if previous_data else None)
            results[key] = {
                "label": label,
                "value": None,
                "previous": prev_val,
                "direction": "▬",
                "unit": "% YoY",
                "source": "FRED",
                "category": "Prices & Inflation",
                "value_date": None,
                "previous_date": None,
                "history": [],
            }

    # Unemployment ────────────────────────────────────────────────────────────
    s = _get("UNRATE")
    cur = safe_last(s)
    prev = safe_prev(s)
    results["UNEMPLOYMENT"] = {
        "label": "Unemployment Rate (U-3)",
        "value": cur,
        "previous": prev,
        "direction": direction_arrow(cur, prev, threshold=0.05),
        "unit": "%",
        "source": "FRED",
        "category": "Real Economy",
        "value_date": _last_date(s),
        "previous_date": _prev_date(s),
        "history": _series_to_history(s),
    }

    # Initial Claims ──────────────────────────────────────────────────────────
    s = _get("ICSA")
    cur = safe_last(s)
    prev = safe_prev(s)
    results["INITIAL_CLAIMS"] = {
        "label": "Initial Jobless Claims",
        "value": cur,
        "previous": prev,
        "direction": direction_arrow(cur, prev, threshold=1000),
        "unit": "K",
        "value_str": f"{cur/1000:.0f}K" if cur else "N/A",
        "source": "FRED",
        "category": "Real Economy",
        "value_date": _last_date(s),
        "previous_date": _prev_date(s),
        "history": _series_to_history(s),
    }

    # NFP MoM change ──────────────────────────────────────────────────────────
    s = _get("PAYEMS")
    nfp_history = []
    nfp_vdate = None
    nfp_pdate = None
    if s is not None and len(s.dropna()) >= 2:
        sd = s.dropna()
        cur_raw = float(sd.iloc[-1])
        prev_raw = float(sd.iloc[-2]) if len(sd) >= 2 else None
        nfp_change = (cur_raw - prev_raw) if prev_raw else None
        prev_change = ((prev_raw - float(sd.iloc[-3]))
                       if len(sd) >= 3 else None)
        nfp_vdate = sd.index[-1].strftime("%Y-%m-%d")
        nfp_pdate = sd.index[-2].strftime("%Y-%m-%d") if len(sd) >= 2 else None
        # Compute MoM change history
        mom = sd.diff().dropna()
        cutoff = mom.index[-1] - timedelta(days=365) if len(mom) > 0 else None
        if cutoff is not None:
            mom = mom[mom.index >= cutoff]
        nfp_history = [[d.strftime("%Y-%m-%d"), round(float(v), 0)]
                       for d, v in mom.items()]
    else:
        nfp_change, prev_change = None, None
    results["NFP"] = {
        "label": "Nonfarm Payrolls (MoM)",
        "value": nfp_change,
        "previous": prev_change,
        "direction": direction_arrow(nfp_change, prev_change, threshold=10),
        "unit": "K",
        "value_str": f"{nfp_change:+.0f}K" if nfp_change else "N/A",
        "source": "FRED",
        "category": "Real Economy",
        "value_date": nfp_vdate,
        "previous_date": nfp_pdate,
        "history": nfp_history,
    }

    # ISM Manufacturing ───────────────────────────────────────────────────────
    # No reliable FRED series exists for the ISM PMI composite (NAPM
    # discontinued, MANEMP is manufacturing employment in thousands).
    # Treat as manual-input indicator with previous-run persistence,
    # same approach as Put/Call Ratio.
    if previous_data and isinstance(previous_data.get("ISM_MFG"), dict):
        prev_ism = previous_data["ISM_MFG"]
        ism_cur = prev_ism.get("value")
        ism_prev = prev_ism.get("previous")
        ism_dir = prev_ism.get("direction", "▬")
        ism_hist = prev_ism.get("history", [])
        ism_vd = prev_ism.get("value_date")
        ism_pd = prev_ism.get("previous_date")
    else:
        ism_cur = ism_prev = None
        ism_dir = "▬"
        ism_hist = []
        ism_vd = ism_pd = None
    results["ISM_MFG"] = {
        "label": "ISM Mfg PMI",
        "value": ism_cur,
        "previous": ism_prev,
        "direction": ism_dir,
        "unit": "",
        "source": "Manual",
        "category": "Real Economy",
        "value_date": ism_vd,
        "previous_date": ism_pd,
        "history": ism_hist,
        "manual_input": True,
    }

    # Consumer Confidence (Michigan) ──────────────────────────────────────────
    s = _get("UMCSENT")
    cur = safe_last(s)
    prev = safe_prev(s)
    results["CONSUMER_CONF"] = {
        "label": "Consumer Confidence (Michigan)",
        "value": cur,
        "previous": prev,
        "direction": direction_arrow(cur, prev, threshold=0.5),
        "unit": "",
        "source": "FRED",
        "category": "Real Economy",
        "value_date": _last_date(s),
        "previous_date": _prev_date(s),
        "history": _series_to_history(s),
    }

    # Credit Spreads ──────────────────────────────────────────────────────────
    # FRED series return values in percentage points (e.g. 1.18 = 118 bps).
    # Convert to basis points by multiplying by 100.
    for key, sid, label in [
        ("IG_SPREAD", "BAMLC0A0CM",  "IG Credit Spread"),
        ("HY_SPREAD", "BAMLH0A0HYM2", "HY Credit Spread"),
    ]:
        s = _get(sid)
        cur_raw = safe_last(s)
        prev_raw = safe_prev(s)
        cur = round(cur_raw * 100, 1) if cur_raw is not None else None
        prev = round(prev_raw * 100, 1) if prev_raw is not None else None
        # Convert history to bps too
        spread_history = []
        if s is not None:
            sd = s.dropna()
            cutoff = sd.index[-1] - timedelta(days=365) if len(sd) > 0 else None
            if cutoff is not None:
                sd = sd[sd.index >= cutoff]
            spread_history = [[d.strftime("%Y-%m-%d"), round(float(v) * 100, 1)]
                              for d, v in sd.items()]
        results[key] = {
            "label": label,
            "value": cur,
            "previous": prev,
            "direction": direction_arrow(cur, prev, threshold=5),
            "unit": "bps",
            "source": "FRED",
            "category": "Market Sentiment",
            "value_date": _last_date(s),
            "previous_date": _prev_date(s),
            "history": spread_history,
        }

    return results


# ─────────────────────────────────────────────────────────────────────────────
# 3. YFINANCE DATA FETCH
# ─────────────────────────────────────────────────────────────────────────────

def fetch_yfinance_data(previous_data: dict) -> dict:
    """Fetch yfinance tickers; return dict of indicator dicts."""
    results = {}

    for key, ticker in config.YFINANCE_TICKERS.items():
        try:
            t = yf.Ticker(ticker)
            hist = t.history(period="1y", auto_adjust=True)
            if hist is None or hist.empty:
                raise ValueError("empty history")
            close = hist["Close"].dropna()
            # Use live price for cur; last daily close as prev (for direction arrow)
            try:
                live = t.fast_info.last_price
                cur = float(live) if live else float(close.iloc[-1])
            except Exception:
                cur = float(close.iloc[-1])
            prev = float(close.iloc[-1]) if len(close) >= 1 else cur
        except Exception as e:
            print(f"  ⚠  yfinance {ticker}: {e}")
            prev_entry = previous_data.get(key, {}) if previous_data else {}
            cur = prev_entry.get("value")
            prev = prev_entry.get("previous")
            close = None
            hist = None

        # Compute SMAs for equities
        sma50 = sma200 = high52w = low52w = None
        if close is not None and len(close) >= 50:
            sma50 = float(close.rolling(50).mean().dropna().iloc[-1])
        if close is not None and len(close) >= 200:
            sma200 = float(close.rolling(200).mean().dropna().iloc[-1])
        if close is not None:
            high52w = float(close.max())
            low52w = float(close.min())

        category_map = {
            "SP500":   "Market Indices",
            "DOW":     "Market Indices",
            "NASDAQ":  "Market Indices",
            "FTSE100": "Market Indices",
            "NIKKEI":  "Market Indices",
            "VIX":     "Market Sentiment",
            "OIL_WTI": "Prices & Inflation",
            "OIL_BRT": "Prices & Inflation",
            "GOLD":    "Prices & Inflation",
            "COPPER":  "Prices & Inflation",
            "DXY":     "Market Sentiment",
        }
        label_map = {
            "SP500":   "S&P 500",
            "DOW":     "Dow Jones",
            "NASDAQ":  "Nasdaq Composite",
            "FTSE100": "FTSE 100",
            "NIKKEI":  "Nikkei 225",
            "VIX":     "VIX",
            "OIL_WTI": "WTI Crude Oil",
            "OIL_BRT": "Brent Crude",
            "GOLD":    "Gold",
            "COPPER":  "Copper",
            "DXY":     "DXY (Dollar Index)",
        }

        # Build history from close series
        yf_history = []
        yf_vdate = None
        yf_pdate = None
        if close is not None and len(close) > 0:
            yf_history = [[d.strftime("%Y-%m-%d"), round(float(v), 2)]
                          for d, v in close.items()]
            yf_vdate = close.index[-1].strftime("%Y-%m-%d")
            yf_pdate = (close.index[-2].strftime("%Y-%m-%d")
                        if len(close) >= 2 else yf_vdate)

        entry = {
            "label":    label_map.get(key, key),
            "value":    cur,
            "previous": prev,
            "direction": direction_arrow(cur, prev,
                         threshold=cur * 0.001 if cur else 0),
            "unit":     "",
            "source":   "yfinance",
            "category": category_map.get(key, "Market Indices"),
            "value_date": yf_vdate,
            "previous_date": yf_pdate,
            "history": yf_history,
        }
        if sma50 is not None:
            entry["sma50"] = sma50
        if sma200 is not None:
            entry["sma200"] = sma200
        if high52w is not None:
            entry["high52w"] = high52w
        if low52w is not None:
            entry["low52w"] = low52w

        # Weekly % change for GOLD signal
        if key == "GOLD" and close is not None and len(close) >= 6:
            entry["weekly_pct"] = float(
                (close.iloc[-1] / close.iloc[-6] - 1) * 100)

        results[key] = entry

    # Put/Call ratio: no reliable free API; mark as manual
    results["PUT_CALL"] = {
        "label":    "Put/Call Ratio",
        "value":    None,
        "previous": None,
        "direction": "▬",
        "unit":     "",
        "source":   "manual",
        "category": "Market Sentiment",
        "manual_input": True,
        "value_date": None,
        "previous_date": None,
        "history": [],
    }

    return results


# ─────────────────────────────────────────────────────────────────────────────
# 4. SIGNAL ASSIGNMENT
# ─────────────────────────────────────────────────────────────────────────────

def assign_signals(indicators: dict, rules: dict) -> dict:
    """Add 'signal' key (BULLISH / BEARISH / NEUTRAL) to each indicator."""

    def signal(key, ind):
        r = rules.get(key, {})
        v = ind.get("value")
        prev = ind.get("previous")

        if v is None:
            return "NEUTRAL"

        # ── VIX ──────────────────────────────────────────────────────────────
        if key == "VIX":
            if v > r.get("bearish_above", 20):
                return "BEARISH"
            if v < r.get("bullish_below", 15):
                return "BULLISH"
            return "NEUTRAL"

        # ── Rates / direction + level based ─────────────────────────────────────
        if key in ("US_10Y", "US_2Y", "US_30Y"):
            if v is None or prev is None:
                return "NEUTRAL"
            change_bps = (v - prev) * 100
            thr = r.get("neutral_change_bps", 3)
            # Level-based: elevated yields are bearish for risk assets
            if key == "US_10Y" and v > 4.5:
                return "BEARISH"
            if key == "US_30Y" and v > 4.8:
                return "BEARISH"
            if change_bps > thr:
                return "BEARISH"
            if change_bps < -thr:
                return "BULLISH"
            return "NEUTRAL"

        # ── Inflation series ──────────────────────────────────────────────────
        if key in ("CPI_YOY", "CORE_CPI_YOY", "CORE_PCE_YOY"):
            if v > r.get("bearish_above", 3.0):
                return "BEARISH"
            if v < r.get("bullish_below", 2.5):
                return "BULLISH"
            return "NEUTRAL"

        # ── Unemployment ──────────────────────────────────────────────────────
        if key == "UNEMPLOYMENT":
            if prev and (v - prev) >= r.get("bearish_rising_by", 0.1):
                return "BEARISH"
            if v > r.get("bearish_above", 4.5):
                return "BEARISH"
            if v < r.get("bullish_below", 4.0):
                return "BULLISH"
            return "NEUTRAL"

        # ── NFP (value is MoM change in thousands) ────────────────────────────
        if key == "NFP":
            if v is None:
                return "NEUTRAL"
            if v < r.get("bearish_below", 0):
                return "BEARISH"
            if v > r.get("bullish_above", 150):
                return "BULLISH"
            return "NEUTRAL"

        # ── ISM Mfg ───────────────────────────────────────────────────────────
        if key == "ISM_MFG":
            if v < r.get("bearish_below", 50):
                return "BEARISH"
            if v > r.get("bullish_above", 55):
                return "BULLISH"
            return "NEUTRAL"

        # ── Credit Spreads (values now in bps after conversion) ─────────────
        if key == "IG_SPREAD":
            if v > r.get("bearish_above", 2.0) * 100:      # 200bps
                return "BEARISH"
            if v < r.get("bullish_below", 1.0) * 100:       # 100bps
                return "BULLISH"
            return "NEUTRAL"

        if key == "HY_SPREAD":
            if v > r.get("bearish_above", 5.0) * 100:       # 500bps
                return "BEARISH"
            if v < r.get("bullish_below", 3.0) * 100:        # 300bps
                return "BULLISH"
            return "NEUTRAL"

        # ── Consumer Confidence ───────────────────────────────────────────────
        if key == "CONSUMER_CONF":
            drop = (prev - v) if prev else 0
            if v < r.get("bearish_below", 80) or drop > r.get("bearish_drop", 5):
                return "BEARISH"
            if v > r.get("bullish_above", 100):
                return "BULLISH"
            return "NEUTRAL"

        # ── Oil ───────────────────────────────────────────────────────────────
        if key == "OIL_WTI":
            if v > r.get("bearish_above", 90):
                return "BEARISH"
            lo = r.get("bullish_low", 50)
            hi = r.get("bullish_high", 75)
            if lo <= v <= hi:
                return "BULLISH"
            return "NEUTRAL"

        if key == "OIL_BRT":
            if v > r.get("bearish_above", 95):
                return "BEARISH"
            lo = r.get("bullish_low", 55)
            hi = r.get("bullish_high", 80)
            if lo <= v <= hi:
                return "BULLISH"
            return "NEUTRAL"

        # ── Copper ───────────────────────────────────────────────────────────
        if key == "COPPER":
            if v > r.get("bullish_above", 4.0):
                return "BULLISH"
            if v < r.get("bearish_below", 3.0):
                return "BEARISH"
            return "NEUTRAL"

        # ── DXY ───────────────────────────────────────────────────────────────
        if key == "DXY":
            if v > r.get("bearish_for_equities_above", 105):
                return "BEARISH"
            if v < r.get("bullish_for_equities_below", 100):
                return "BULLISH"
            return "NEUTRAL"

        # ── Initial Claims ───────────────────────────────────────────────────
        if key == "INITIAL_CLAIMS":
            if v > r.get("bearish_above", 300000):
                return "BEARISH"
            if v < r.get("bullish_below", 220000):
                return "BULLISH"
            return "NEUTRAL"

        # ── Equity Indices (SMA-based) ───────────────────────────────────────
        if key in ("SP500", "DOW", "NASDAQ", "FTSE100", "NIKKEI"):
            sma200 = ind.get("sma200")
            sma50  = ind.get("sma50")
            if sma200 and v < sma200:
                return "BEARISH"
            if sma50 and v > sma50:
                return "BULLISH"
            return "NEUTRAL"

        # ── Gold ──────────────────────────────────────────────────────────────
        if key == "GOLD":
            w = ind.get("weekly_pct", 0) or 0
            if w < r.get("bearish_weekly_pct_below", -3.0):
                return "BEARISH"
            if w > r.get("bullish_weekly_pct_above", 1.5):
                return "BULLISH"
            return "NEUTRAL"

        # ── 2s10s Spread ──────────────────────────────────────────────────────
        if key == "SPREAD_2S10S":
            if v < r.get("bearish_below", 0):
                return "BEARISH"
            if v > r.get("bullish_above", 0.5):
                return "BULLISH"
            return "NEUTRAL"

        # ── Fed Funds (direction) ─────────────────────────────────────────────
        if key == "FED_FUNDS":
            if v is None or prev is None:
                return "NEUTRAL"
            if v < prev - 0.01:
                return "BULLISH"   # cut
            if v > prev + 0.01:
                return "BEARISH"   # hike
            return "NEUTRAL"

        # ── Put/Call ──────────────────────────────────────────────────────────
        if key == "PUT_CALL":
            if v is None:
                return "NEUTRAL"
            if v > r.get("contrarian_bullish_above", 1.2):
                return "BULLISH"  # contrarian
            if v > r.get("bearish_above", 1.0):
                return "BEARISH"
            if v < r.get("bullish_below", 0.7):
                return "BULLISH"
            return "NEUTRAL"

        return "NEUTRAL"

    for key, ind in indicators.items():
        ind["signal"] = signal(key, ind)

    return indicators


# ─────────────────────────────────────────────────────────────────────────────
# 5. COMMENTARY TEMPLATES
# ─────────────────────────────────────────────────────────────────────────────

def generate_commentary(key: str, ind: dict) -> str:
    v = ind.get("value")
    prev = ind.get("previous")
    sig = ind.get("signal", "NEUTRAL")
    direction = ind.get("direction", "▬")
    change = (v - prev) if (v is not None and prev is not None) else None

    if v is None:
        return "Data unavailable — using last known value."

    def chg_str(val, suffix="", decimals=2, scale=1):
        if val is None:
            return ""
        return f"{val * scale:+.{decimals}f}{suffix}"

    if key == "VIX":
        base = f"VIX at {v:.2f} {direction}."
        if v > 40:
            return base + " Extreme fear — potential capitulation territory."
        if v > 30:
            return base + " Deep fear zone; risk-off sentiment dominant."
        if v > 20:
            return base + " Elevated anxiety; above 20 = fear zone."
        if v < 15:
            return base + " Complacency zone; historically low volatility."
        return base + " Moderate concern; market in transition."

    if key == "FED_FUNDS":
        lo = ind.get("value_str", f"{v:.2f}%")
        if change and change < -0.01:
            return f"Fed Funds at {lo} — cut cycle underway. Accommodative pivot in progress."
        if change and change > 0.01:
            return f"Fed Funds at {lo} — hiking cycle ongoing. Tighter financial conditions."
        return f"Fed Funds at {lo} — on hold. Market watching for next pivot signal."

    if key == "US_10Y":
        bps = change * 100 if change else 0
        base = f"10Y yield at {v:.2f}% {direction} ({bps:+.0f}bps)."
        if bps > 20:
            return base + " Sharp move higher — tightening financial conditions."
        if bps < -20:
            return base + " Significant rally — flight-to-safety or dovish re-pricing."
        if v > 4.5:
            return base + " Yield elevated; pressure on valuations and credit."
        return base + " Yield within recent range."

    if key == "US_2Y":
        bps = change * 100 if change else 0
        return (f"2Y yield at {v:.2f}% {direction} ({bps:+.0f}bps). "
                f"{'High — pricing in prolonged Fed hold.' if v > 4.5 else 'Reflecting current rate expectations.'}")

    if key == "US_30Y":
        bps = change * 100 if change else 0
        return (f"30Y yield at {v:.2f}% {direction} ({bps:+.0f}bps). "
                f"{'Long-end elevated; term premium rising.' if v > 4.5 else 'Long-end contained.'}")

    if key == "SPREAD_2S10S":
        base = f"2s10s spread at {v:.2f}% {direction}."
        if v < 0:
            return base + " Curve inverted — historically precedes recession."
        if v > 0.5:
            return base + " Curve steepening — growth optimism or reflation trade."
        return base + " Flat curve; economic uncertainty persists."

    if key == "CPI_YOY":
        base = f"CPI at {v:.1f}% YoY {direction}."
        if v > 4:
            return base + " Well above Fed target; strong case for higher-for-longer."
        if v > 3:
            return base + " Above Fed target; disinflation progress stalling."
        if v > 2.5:
            return base + " Approaching target but sticky; last-mile inflation challenge."
        if v <= 2.0:
            return base + " At or below target; opens door to Fed cuts."
        return base + " Near Fed 2% target."

    if key == "CORE_CPI_YOY":
        base = f"Core CPI at {v:.1f}% YoY {direction}."
        if v > 3.5:
            return base + " Services inflation sticky; Fed maintaining restrictive stance."
        if v > 2.5:
            return base + " Elevated core; above Fed comfort zone."
        return base + " Core inflation cooling — supportive of easing."

    if key == "CORE_PCE_YOY":
        base = f"Core PCE at {v:.1f}% YoY {direction}."
        if v > 2.5:
            return base + " Fed's preferred gauge above target — limits easing."
        return base + " Fed's preferred gauge near target."

    if key == "UNEMPLOYMENT":
        base = f"Unemployment at {v:.1f}% {direction}."
        if change and change >= 0.2:
            return base + f" Rose {change:+.1f}pp — Sahm Rule territory; recession risk elevated."
        if v < 4.0:
            return base + " Sub-4% unemployment; labor market tight."
        if v > 5.0:
            return base + " Softening labor market; recession concerns mounting."
        return base + " Labor market normalizing."

    if key == "INITIAL_CLAIMS":
        vs = ind.get("value_str", f"{v/1000:.0f}K")
        base = f"Initial claims at {vs} {direction}."
        if v > 300000:
            return base + " Above 300K signals meaningful labor market deterioration."
        if v < 220000:
            return base + " Sub-220K — historically tight labor market."
        return base + " Within normal range."

    if key == "NFP":
        vs = ind.get("value_str", f"{v:+.0f}K")
        base = f"NFP at {vs} {direction}."
        nfp_k = v
        if nfp_k < 0:
            return base + " Negative prints — labor market contraction signal."
        if nfp_k < 100:
            return base + " Below consensus — cooling labor market."
        if nfp_k > 300:
            return base + " Blowout number — hot labor market, Fed staying hawkish."
        return base + " Solid print; labor market resilient."

    if key == "ISM_MFG":
        base = f"ISM Mfg at {v:.1f} {direction}."
        if v < 45:
            return base + " Deep contraction; manufacturing in recession."
        if v < 50:
            return base + " Below 50 = contraction in manufacturing sector."
        if v > 55:
            return base + " Above 55 — strong expansion; reflation signal."
        return base + " Modest expansion above key 50 threshold."

    if key == "CONSUMER_CONF":
        base = f"Consumer confidence at {v:.1f} {direction}."
        drop = (prev - v) if prev else 0
        if drop > 5:
            return base + f" Fell {drop:.1f}pts — deteriorating consumer sentiment."
        if v < 80:
            return base + " Below 80 — significant pessimism; spending headwinds."
        if v > 100:
            return base + " Above 100 — healthy confidence; supports consumption."
        return base + " Moderate confidence."

    if key == "IG_SPREAD":
        base = f"IG spread at {v:.0f}bps {direction}."
        if v > 200:
            return base + " Elevated — credit markets pricing stress."
        if v < 100:
            return base + " Tight — benign credit environment."
        return base + " Within normal range."

    if key == "HY_SPREAD":
        base = f"HY spread at {v:.0f}bps {direction}."
        if v > 700:
            return base + " Extreme stress — approaching GFC-level credit fear."
        if v > 500:
            return base + " Above 500bps — junk credit distress; risk-off."
        if v < 300:
            return base + " Below 300bps — compressed; complacency or strong growth."
        return base + " Moderate risk premium."

    if key == "SP500":
        sma50 = ind.get("sma50")
        sma200 = ind.get("sma200")
        high52w = ind.get("high52w")
        base = f"S&P 500 at {v:,.0f} {direction}."
        parts = [base]
        if sma200 and v < sma200:
            parts.append(f"Below 200-day SMA ({sma200:,.0f}) — trend broken.")
        elif sma50 and v > sma50:
            parts.append(f"Above 50-day SMA ({sma50:,.0f}) — uptrend intact.")
        if high52w and v < high52w:
            pct_off = (v / high52w - 1) * 100
            parts.append(f"{pct_off:.1f}% from 52-week high.")
        return " ".join(parts)

    if key == "DOW":
        return f"Dow Jones at {v:,.0f} {direction}."

    if key == "NASDAQ":
        sma200 = ind.get("sma200")
        base = f"Nasdaq at {v:,.0f} {direction}."
        if sma200 and v < sma200:
            return base + " Below 200-day SMA — tech trend under pressure."
        return base + " Tech complex tracking broader risk tone."

    if key == "FTSE100":
        return f"FTSE 100 at {v:,.0f} {direction}. UK equities reflecting global risk sentiment."

    if key == "NIKKEI":
        return f"Nikkei at {v:,.0f} {direction}. JPY dynamics and global tech flows key drivers."

    if key == "OIL_WTI":
        base = f"WTI at ${v:.2f} {direction}."
        if v > 100:
            return base + " Triple digits — supply shock; stagflation risk elevated."
        if v > 90:
            return base + " Elevated — headwind for inflation and consumer spending."
        if v < 50:
            return base + " Low oil — demand destruction signal or oversupply."
        if 50 <= v <= 75:
            return base + " Goldilocks zone — supportive without fueling inflation."
        return base + " Moderate level; watching OPEC supply signals."

    if key == "OIL_BRT":
        return f"Brent at ${v:.2f} {direction}. Global crude benchmark tracking geopolitical risk."

    if key == "GOLD":
        w = ind.get("weekly_pct", 0) or 0
        base = f"Gold at ${v:,.0f} {direction} ({w:+.1f}% wk)."
        if w < -3:
            return base + " Sharp weekly drop — forced liquidation or dollar strength."
        if w > 3:
            return base + " Strong weekly gain — inflation hedge / safe haven bid."
        if v > 2800:
            return base + " All-time high territory — persistent safe-haven demand."
        return base + " Range-bound; watching real rates and dollar direction."

    if key == "COPPER":
        return (f"Copper at ${v:.3f} {direction}. "
                f"{'Above $4 — strong global growth signal.' if v > 4 else 'Watching China demand and global PMI.'}")

    if key == "DXY":
        base = f"DXY at {v:.2f} {direction}."
        if v > 105:
            return base + " Strong dollar — headwind for multinational earnings and EM."
        if v < 100:
            return base + " Weak dollar — supports commodities and overseas earnings."
        return base + " Dollar neutral; tracking rate differential."

    if key == "PUT_CALL":
        if v is None:
            return "Put/Call ratio — manual input required. Update via Edit Mode."
        if v > 1.2:
            return f"Put/Call at {v:.2f} — extreme fear; contrarian BULLISH signal."
        if v > 1.0:
            return f"Put/Call at {v:.2f} — fear elevated; bearish sentiment."
        if v < 0.7:
            return f"Put/Call at {v:.2f} — complacency / bullish positioning."
        return f"Put/Call at {v:.2f} — neutral sentiment."

    # Generic fallback
    vs = f"{v:.2f}" if isinstance(v, float) else str(v)
    return f"{ind.get('label', key)} at {vs} {direction}. Signal: {sig}."


# ─────────────────────────────────────────────────────────────────────────────
# 5b. MACRO NARRATIVE
# ─────────────────────────────────────────────────────────────────────────────

def generate_narrative(indicators: dict) -> str:
    """Generate 2-3 sentence auto-generated macro narrative."""
    vix_v = indicators.get("VIX", {}).get("value")
    oil_v = indicators.get("OIL_WTI", {}).get("value")
    sp_v = indicators.get("SP500", {}).get("value")
    sp_sma200 = indicators.get("SP500", {}).get("sma200")
    sp_high = indicators.get("SP500", {}).get("high52w")
    unemp_v = indicators.get("UNEMPLOYMENT", {}).get("value")
    unemp_prev = indicators.get("UNEMPLOYMENT", {}).get("previous")
    nfp_v = indicators.get("NFP", {}).get("value")
    cpi_v = indicators.get("CPI_YOY", {}).get("value")
    pce_v = indicators.get("CORE_PCE_YOY", {}).get("value")
    ff = indicators.get("FED_FUNDS", {})
    ff_str = ff.get("value_str", "N/A")
    hy_v = indicators.get("HY_SPREAD", {}).get("value")
    spread_v = indicators.get("SPREAD_2S10S", {}).get("value")
    consumer_v = indicators.get("CONSUMER_CONF", {}).get("value")
    dxy_v = indicators.get("DXY", {}).get("value")

    bull = sum(1 for i in indicators.values() if i.get("signal") == "BULLISH")
    bear = sum(1 for i in indicators.values() if i.get("signal") == "BEARISH")

    sentences = []

    # ── Sentence 1: Dominant macro theme ─────────────────────────────────────
    inflation_hot = (oil_v and oil_v > 85) or (cpi_v and cpi_v > 3.0) or (pce_v and pce_v > 2.8)
    labor_weak = (nfp_v is not None and nfp_v < 50) or (
        unemp_v and unemp_prev and (unemp_v - unemp_prev) >= 0.1)
    risk_off = vix_v and vix_v > 22 and sp_sma200 and sp_v and sp_v < sp_sma200

    if inflation_hot and labor_weak:
        inf_bits = []
        if oil_v and oil_v > 85:
            inf_bits.append(f"oil above ${oil_v:.0f}")
        if cpi_v and cpi_v > 3.0:
            inf_bits.append(f"CPI at {cpi_v:.1f}%")
        elif pce_v and pce_v > 2.8:
            inf_bits.append(f"Core PCE at {pce_v:.1f}%")
        lab_bits = []
        if nfp_v is not None and nfp_v < 50:
            lab_bits.append(f"NFP at {nfp_v:+.0f}K")
        if unemp_v and unemp_prev and (unemp_v - unemp_prev) >= 0.1:
            lab_bits.append(f"unemployment rising to {unemp_v:.1f}%")
        sentences.append(
            f"Stagflation risk elevated: {' and '.join(inf_bits)}"
            f" while {' and '.join(lab_bits)}.")
    elif risk_off:
        drawdown = ((sp_v / sp_high - 1) * 100) if sp_high else 0
        s = "Risk-off regime dominant: S&P 500 below 200-day SMA"
        if drawdown < -3:
            s += f" ({drawdown:.1f}% from 52-week high)"
        s += f" with VIX at {vix_v:.0f}."
        sentences.append(s)
    elif bear > bull + 4:
        sentences.append(
            f"Broad macro deterioration \u2014 {bear} bearish vs {bull} bullish indicators.")
    elif bull > bear + 4:
        sentences.append(
            f"Constructive macro backdrop with {bull} bullish signals leading.")
    else:
        sentences.append(
            f"Mixed macro regime \u2014 {bear} bearish, {bull} bullish signals competing.")

    # ── Sentence 2: Fed / rates context ──────────────────────────────────────
    s2 = f"Fed on hold at {ff_str}"
    if ff.get("value") and ff.get("previous"):
        if ff["value"] < ff["previous"] - 0.01:
            s2 = f"Fed cutting \u2014 funds rate at {ff_str}"
        elif ff["value"] > ff["previous"] + 0.01:
            s2 = f"Fed hiking \u2014 funds rate at {ff_str}"
    if spread_v is not None:
        if spread_v < 0:
            s2 += f"; yield curve inverted ({spread_v:+.2f}%) signaling recession risk"
        elif spread_v > 0.3:
            s2 += f"; curve steepening at {spread_v:.2f}%"
    s2 += "."
    sentences.append(s2)

    # ── Sentence 3: Financial conditions / market stress ─────────────────────
    stress = []
    if vix_v and vix_v > 25:
        stress.append(f"elevated volatility (VIX {vix_v:.0f})")
    if hy_v and hy_v > 400:
        stress.append(f"HY spreads at {hy_v:.0f}bps")
    if consumer_v and consumer_v < 70:
        stress.append(f"consumer sentiment depressed ({consumer_v:.0f})")
    if dxy_v and dxy_v < 98:
        stress.append(f"weak dollar (DXY {dxy_v:.1f})")
    elif dxy_v and dxy_v > 106:
        stress.append(f"strong dollar (DXY {dxy_v:.1f}) pressuring EM")
    if stress:
        tag = ("risk-off positioning dominant" if bear > bull
               else "markets navigating crosscurrents")
        sentences.append(" | ".join(stress[:3]) + f" \u2014 {tag}.")

    return " ".join(sentences)


# ─────────────────────────────────────────────────────────────────────────────
# 5c. UPCOMING RELEASES
# ─────────────────────────────────────────────────────────────────────────────

def get_upcoming_releases(today=None):
    """Return list of (name, date_str, days_away) for next 5 major releases."""
    from datetime import date as date_cls

    if today is None:
        today = datetime.now(timezone.utc).date()

    def next_weekday_on_or_after(d, wd):
        return d + timedelta(days=(wd - d.weekday()) % 7)

    def first_weekday_of_month(y, m, wd):
        return next_weekday_on_or_after(date_cls(y, m, 1), wd)

    def nm(y, m):
        return (y + 1, 1) if m == 12 else (y, m + 1)

    releases = []

    # NFP — first Friday of month
    nfp = first_weekday_of_month(today.year, today.month, 4)
    if nfp <= today:
        y, m = nm(today.year, today.month)
        nfp = first_weekday_of_month(y, m, 4)
    releases.append(("Nonfarm Payrolls", nfp, "Labour"))

    # CPI — approx second Wednesday of month
    cpi = first_weekday_of_month(today.year, today.month, 2) + timedelta(days=7)
    if cpi <= today:
        y, m = nm(today.year, today.month)
        cpi = first_weekday_of_month(y, m, 2) + timedelta(days=7)
    releases.append(("CPI Release", cpi, "Inflation"))

    # PCE — last Friday of month
    def last_friday(y, m):
        ny, nm_ = nm(y, m)
        last = date_cls(ny, nm_, 1) - timedelta(days=1)
        return last - timedelta(days=(last.weekday() - 4) % 7)
    pce = last_friday(today.year, today.month)
    if pce <= today:
        y, m = nm(today.year, today.month)
        pce = last_friday(y, m)
    releases.append(("Core PCE / Personal Income", pce, "Inflation"))

    # Initial Claims — next Thursday
    nxt_thu = today + timedelta(days=((3 - today.weekday()) % 7) or 7)
    releases.append(("Initial Jobless Claims", nxt_thu, "Labour"))

    # ISM Manufacturing — first business day of month
    def first_bday(y, m):
        d = date_cls(y, m, 1)
        while d.weekday() >= 5:
            d += timedelta(days=1)
        return d
    ism = first_bday(today.year, today.month)
    if ism <= today:
        y, m = nm(today.year, today.month)
        ism = first_bday(y, m)
    releases.append(("ISM Manufacturing PMI", ism, "Activity"))

    # FOMC — hardcoded schedule
    fomc_dates = [
        date_cls(2025, 1, 29), date_cls(2025, 3, 19), date_cls(2025, 5, 7),
        date_cls(2025, 6, 18), date_cls(2025, 7, 30), date_cls(2025, 9, 17),
        date_cls(2025, 10, 29), date_cls(2025, 12, 17),
        date_cls(2026, 1, 28), date_cls(2026, 3, 18), date_cls(2026, 4, 29),
        date_cls(2026, 6, 17), date_cls(2026, 7, 29), date_cls(2026, 9, 16),
        date_cls(2026, 10, 28), date_cls(2026, 12, 16),
        date_cls(2027, 1, 27), date_cls(2027, 3, 17),
    ]
    for d in fomc_dates:
        if d > today:
            releases.append(("FOMC Decision", d, "Rates"))
            break

    # Michigan Consumer Sentiment — ~2nd Friday of month
    mich = first_weekday_of_month(today.year, today.month, 4) + timedelta(days=7)
    if mich <= today:
        y, m = nm(today.year, today.month)
        mich = first_weekday_of_month(y, m, 4) + timedelta(days=7)
    releases.append(("Michigan Consumer Sentiment", mich, "Sentiment"))

    releases.sort(key=lambda x: x[1])
    result = []
    for name, d, cat in releases:
        days = (d - today).days
        result.append((name, d.strftime("%-d %b"), days, cat))
    return result[:5]


# ─────────────────────────────────────────────────────────────────────────────
# 5d. CORRELATION MAP
# ─────────────────────────────────────────────────────────────────────────────

CORRELATIONS = {
    "FED_FUNDS":    ["US_2Y", "US_10Y", "SPREAD_2S10S"],
    "US_2Y":        ["FED_FUNDS", "US_10Y", "SPREAD_2S10S"],
    "US_10Y":       ["US_2Y", "US_30Y", "SP500"],
    "US_30Y":       ["US_10Y", "SPREAD_2S10S", "GOLD"],
    "SPREAD_2S10S": ["US_2Y", "US_10Y", "FED_FUNDS"],
    "CPI_YOY":      ["CORE_CPI_YOY", "OIL_WTI", "FED_FUNDS"],
    "CORE_CPI_YOY": ["CPI_YOY", "CORE_PCE_YOY", "FED_FUNDS"],
    "CORE_PCE_YOY": ["CPI_YOY", "CORE_CPI_YOY", "FED_FUNDS"],
    "UNEMPLOYMENT":  ["NFP", "INITIAL_CLAIMS", "CONSUMER_CONF"],
    "INITIAL_CLAIMS":["UNEMPLOYMENT", "NFP", "SP500"],
    "NFP":          ["UNEMPLOYMENT", "INITIAL_CLAIMS", "SP500"],
    "ISM_MFG":      ["COPPER", "SP500", "CONSUMER_CONF"],
    "CONSUMER_CONF":["SP500", "UNEMPLOYMENT", "NFP"],
    "IG_SPREAD":    ["HY_SPREAD", "VIX", "SP500"],
    "HY_SPREAD":    ["IG_SPREAD", "VIX", "SP500"],
    "VIX":          ["SP500", "HY_SPREAD", "PUT_CALL"],
    "SP500":        ["VIX", "NASDAQ", "US_10Y"],
    "DOW":          ["SP500", "NASDAQ", "VIX"],
    "NASDAQ":       ["SP500", "DOW", "VIX"],
    "FTSE100":      ["SP500", "NIKKEI", "OIL_BRT"],
    "NIKKEI":       ["SP500", "FTSE100", "DXY"],
    "OIL_WTI":      ["CPI_YOY", "OIL_BRT", "SP500"],
    "OIL_BRT":      ["OIL_WTI", "CPI_YOY", "DXY"],
    "GOLD":         ["DXY", "VIX", "US_10Y"],
    "COPPER":       ["ISM_MFG", "SP500", "DXY"],
    "DXY":          ["GOLD", "SP500", "US_10Y"],
    "PUT_CALL":     ["VIX", "SP500", "HY_SPREAD"],
}


# ─────────────────────────────────────────────────────────────────────────────
# 6. HTML GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

SECTION_ORDER = [
    "Rates & Monetary Policy",
    "Real Economy",
    "Prices & Inflation",
    "Market Sentiment",
    "Market Indices",
]

SECTION_ICONS = {
    "Rates & Monetary Policy": "🏦",
    "Real Economy":            "📊",
    "Prices & Inflation":      "📈",
    "Market Sentiment":        "😰",
    "Market Indices":          "💹",
}

def value_display(key: str, ind: dict) -> str:
    """Return a formatted display string for the value."""
    v = ind.get("value")
    vs = ind.get("value_str")
    if vs:
        return vs
    if v is None:
        return "N/A"
    unit = ind.get("unit", "")
    if key in ("SP500", "DOW", "NASDAQ", "FTSE100", "NIKKEI"):
        return f"{v:,.0f}"
    if key in ("GOLD", "OIL_WTI", "OIL_BRT"):
        return f"${v:,.2f}"
    if key in ("COPPER",):
        return f"${v:.3f}"
    if unit == "%" or unit.endswith("% YoY"):
        return f"{v:.2f}%"
    if unit == "bps":
        return f"{v:.0f}bps"
    if key == "DXY":
        return f"{v:.2f}"
    return f"{v:.2f}"


def prev_display(key: str, ind: dict) -> str:
    prev = ind.get("previous")
    if prev is None:
        return "—"
    unit = ind.get("unit", "")
    if key in ("SP500", "DOW", "NASDAQ", "FTSE100", "NIKKEI"):
        return f"{prev:,.0f}"
    if key in ("GOLD", "OIL_WTI", "OIL_BRT"):
        return f"${prev:,.2f}"
    if key == "INITIAL_CLAIMS":
        return f"{prev/1000:.0f}K"
    if key == "NFP":
        return f"{prev:+.0f}K"
    if unit == "%" or unit.endswith("% YoY"):
        return f"{prev:.2f}%"
    if unit == "bps":
        return f"{prev:.0f}bps"
    return f"{prev:.2f}"


def prev_date_display(ind: dict) -> str:
    """Format previous_date as '(18 Mar)' or '' if unavailable."""
    pd = ind.get("previous_date")
    if not pd:
        return ""
    try:
        dt = datetime.strptime(pd, "%Y-%m-%d")
        return f"({dt.strftime('%-d %b')})"
    except Exception:
        return ""


RANGE_BAR_KEYS = {
    "SP500", "DOW", "NASDAQ", "FTSE100", "NIKKEI",
    "VIX", "OIL_WTI", "OIL_BRT", "GOLD", "COPPER", "DXY",
    "US_10Y", "US_2Y", "US_30Y",
}

def range_bar_html(key: str, ind: dict) -> str:
    """Return HTML for a 52-week range bar, or empty string if no data."""
    if key not in RANGE_BAR_KEYS:
        return ""
    hi = ind.get("high52w")
    lo = ind.get("low52w")
    v = ind.get("value")
    if hi is None or lo is None or v is None or hi == lo:
        return ""
    pct = max(0, min(100, (v - lo) / (hi - lo) * 100))
    # Format labels based on indicator type
    unit = ind.get("unit", "")
    if key in ("SP500", "DOW", "NASDAQ", "FTSE100", "NIKKEI"):
        lo_s, hi_s = f"{lo:,.0f}", f"{hi:,.0f}"
    elif key in ("GOLD", "OIL_WTI", "OIL_BRT"):
        lo_s, hi_s = f"${lo:,.0f}", f"${hi:,.0f}"
    elif key == "COPPER":
        lo_s, hi_s = f"${lo:.2f}", f"${hi:.2f}"
    elif unit == "%" or unit.endswith("% YoY"):
        lo_s, hi_s = f"{lo:.2f}%", f"{hi:.2f}%"
    elif unit == "K":
        lo_s, hi_s = f"{lo/1000:.0f}K", f"{hi/1000:.0f}K"
    else:
        lo_s, hi_s = f"{lo:.2f}", f"{hi:.2f}"
    return (f'<div class="range-bar" title="52-week range">'
            f'<div class="range-label">52w range</div>'
            f'<div class="range-track">'
            f'<div class="range-fill" style="width:{pct:.1f}%"></div>'
            f'<div class="range-marker" style="left:{pct:.1f}%"></div>'
            f'</div>'
            f'<div class="range-labels"><span>{lo_s}</span>'
            f'<span>{hi_s}</span></div></div>')


def sparkline_svg(key, ind):
    """Generate inline SVG sparkline (120x30) from last 30 data points."""
    history = ind.get("history", [])
    if not history or len(history) < 2:
        return ""
    pts = history[-30:]
    values = [p[1] for p in pts]
    mn, mx = min(values), max(values)
    rng = mx - mn if mx != mn else 1
    w, h = 120, 30
    pad = 2
    coords = []
    for i, v in enumerate(values):
        x = pad + (w - 2 * pad) * i / (len(values) - 1)
        y = h - pad - (h - 2 * pad) * (v - mn) / rng
        coords.append(f"{x:.1f},{y:.1f}")
    sig = ind.get("signal", "NEUTRAL")
    color = "#00e676" if sig == "BULLISH" else "#ff4d4d" if sig == "BEARISH" else "#ffd24d"
    poly = " ".join(coords)
    fill_poly = f"{pad:.1f},{h - pad:.1f} " + poly + f" {w - pad:.1f},{h - pad:.1f}"
    lx = coords[-1].split(",")[0]
    ly = coords[-1].split(",")[1]
    return (f'<svg class="sparkline" width="{w}" height="{h}" viewBox="0 0 {w} {h}">'
            f'<polyline points="{fill_poly}" fill="{color}22" stroke="none"/>'
            f'<polyline points="{poly}" fill="none" stroke="{color}" '
            f'stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/>'
            f'<circle cx="{lx}" cy="{ly}" r="2" fill="{color}"/>'
            f'</svg>')


def compute_momentum(ind):
    """Return 'ACCEL', 'DECEL', or None based on rate of change analysis."""
    history = ind.get("history", [])
    if len(history) < 6:
        return None
    pts = [h[1] for h in history[-6:]]
    changes = [abs(pts[i + 1] - pts[i]) for i in range(5)]
    latest = changes[-1]
    avg_prior = sum(changes[:-1]) / 4
    if avg_prior == 0:
        return None
    ratio = latest / avg_prior
    if ratio > 1.5:
        return "ACCEL"
    if ratio < 0.5:
        return "DECEL"
    return None


def detect_changes(indicators):
    """Return list of meaningful changes as (key, label, old_str, new_str, context)."""
    relative_keys = {"SP500", "DOW", "NASDAQ", "FTSE100", "NIKKEI",
                     "OIL_WTI", "OIL_BRT", "GOLD", "COPPER", "DXY", "VIX"}
    rate_keys = {"FED_FUNDS", "US_2Y", "US_10Y", "US_30Y", "SPREAD_2S10S"}
    spread_keys = {"IG_SPREAD", "HY_SPREAD"}
    changes = []
    for key, ind in indicators.items():
        v = ind.get("value")
        prev = ind.get("previous")
        if v is None or prev is None:
            continue
        diff = v - prev
        is_significant = False
        context = ""
        if key in relative_keys:
            pct = abs(diff / prev * 100) if prev != 0 else 0
            if pct > 1.0:
                is_significant = True
                context = f"{diff / prev * 100:+.1f}%"
        elif key in rate_keys:
            bps = abs(diff * 100)
            if bps >= 5:
                is_significant = True
                context = f"{diff * 100:+.0f}bps"
        elif key in spread_keys:
            # Already in bps after conversion
            if abs(diff) >= 5:
                is_significant = True
                context = f"{diff:+.0f}bps"
        elif key == "NFP":
            if abs(diff) >= 20:
                is_significant = True
                context = f"{diff:+.0f}K vs prior"
        elif key == "UNEMPLOYMENT":
            if abs(diff) >= 0.1:
                is_significant = True
                context = f"{diff:+.1f}pp"
        elif key == "INITIAL_CLAIMS":
            if abs(diff) >= 5000:
                is_significant = True
                context = f"{diff / 1000:+.0f}K"
        elif key in ("CPI_YOY", "CORE_CPI_YOY", "CORE_PCE_YOY"):
            if abs(diff) >= 0.1:
                is_significant = True
                context = f"{diff:+.1f}pp"
        elif key in ("ISM_MFG", "CONSUMER_CONF"):
            if abs(diff) >= 1.0:
                is_significant = True
                context = f"{diff:+.1f}pts"
        if is_significant:
            changes.append((key, ind.get("label", key),
                           prev_display(key, ind),
                           value_display(key, ind), context))
    return changes


INDICATOR_GLOSSARY = {
    "FED_FUNDS": "The interest rate banks charge each other overnight, set by the Federal Reserve.",
    "US_2Y": "Yield on 2-year US Treasury bonds. Reflects expectations for Fed policy.",
    "US_10Y": "Yield on 10-year Treasuries. Key benchmark for mortgages and corporate bonds.",
    "US_30Y": "Yield on 30-year Treasuries. Reflects long-term inflation expectations.",
    "SPREAD_2S10S": "Difference between 10Y and 2Y yields. Negative (inverted) = recession signal.",
    "CPI_YOY": "Consumer Price Index year-over-year. Measures headline inflation.",
    "CORE_CPI_YOY": "CPI excluding food &amp; energy. Shows underlying inflation trend.",
    "CORE_PCE_YOY": "Personal Consumption Expenditures ex food &amp; energy. Fed&#39;s preferred inflation gauge.",
    "UNEMPLOYMENT": "Percentage of labor force actively seeking work. Key economic health indicator.",
    "INITIAL_CLAIMS": "Weekly new unemployment insurance filings. Early labor market warning signal.",
    "NFP": "Monthly change in US non-farm employment. Most-watched jobs report.",
    "ISM_MFG": "Manufacturing purchasing managers survey. Above 50 = expansion, below = contraction.",
    "CONSUMER_CONF": "University of Michigan household sentiment survey. Drives spending expectations.",
    "IG_SPREAD": "Extra yield on investment-grade corporate bonds over Treasuries. Measures credit risk.",
    "HY_SPREAD": "Extra yield on high-yield (junk) bonds over Treasuries. Spikes during stress.",
    "VIX": "Implied volatility of S&amp;P 500 options. The &#39;fear index&#39; — spikes during panic.",
    "SP500": "Market-cap weighted index of 500 largest US companies. Primary US equity benchmark.",
    "DOW": "Price-weighted index of 30 large US blue-chip stocks.",
    "NASDAQ": "Tech-heavy index. Proxy for growth and technology sentiment.",
    "FTSE100": "100 largest UK-listed companies. UK equity benchmark.",
    "NIKKEI": "225 major Japanese companies. Reflects Asian sentiment and yen dynamics.",
    "OIL_WTI": "West Texas Intermediate crude. US oil benchmark affecting inflation and energy stocks.",
    "OIL_BRT": "Brent crude. Global oil benchmark for international pricing.",
    "GOLD": "Price per ounce. Safe-haven asset rising during inflation and uncertainty.",
    "COPPER": "&#39;Dr. Copper&#39; — demand reflects global industrial activity and growth.",
    "DXY": "US Dollar vs 6 major currencies. Strong dollar hurts exports and emerging markets.",
    "PUT_CALL": "Ratio of put to call options. High = fear; extreme readings are contrarian signals.",
}


# ─────────────────────────────────────────────────────────────────────────────
# 6b. MACRO MAP / NETWORK DIAGRAM
# ─────────────────────────────────────────────────────────────────────────────

def compute_extremity(ind):
    """Return 0-1 extremity score based on 52-week range."""
    hi = ind.get("high52w")
    lo = ind.get("low52w")
    v = ind.get("value")
    if hi is None or lo is None or v is None or hi == lo:
        return 0.5
    mid = (hi + lo) / 2
    rng = (hi - lo) / 2
    return min(1.0, abs(v - mid) / rng) if rng > 0 else 0.5


def build_macro_map_data(indicators):
    """Return (nodes_json, edges_json) for the D3 network diagram."""
    sig_colors = {"BULLISH": "#00e676", "BEARISH": "#ff4d4d", "NEUTRAL": "#ffd24d"}
    short_labels = {
        "FED_FUNDS": "Fed Funds", "US_2Y": "2Y", "US_10Y": "10Y",
        "US_30Y": "30Y", "SPREAD_2S10S": "2s10s",
        "CPI_YOY": "CPI", "CORE_CPI_YOY": "Core CPI", "CORE_PCE_YOY": "Core PCE",
        "UNEMPLOYMENT": "Unemp", "INITIAL_CLAIMS": "Claims", "NFP": "NFP",
        "ISM_MFG": "ISM", "CONSUMER_CONF": "Cons.Conf",
        "IG_SPREAD": "IG", "HY_SPREAD": "HY",
        "VIX": "VIX", "SP500": "S&P", "DOW": "Dow", "NASDAQ": "Nasdaq",
        "FTSE100": "FTSE", "NIKKEI": "Nikkei",
        "OIL_WTI": "WTI", "OIL_BRT": "Brent", "GOLD": "Gold",
        "COPPER": "Copper", "DXY": "DXY", "PUT_CALL": "P/C",
    }
    positions = {
        "FED_FUNDS": (150, 120), "US_2Y": (300, 80), "US_10Y": (450, 100),
        "US_30Y": (450, 200), "SPREAD_2S10S": (600, 80),
        "CPI_YOY": (600, 320), "CORE_CPI_YOY": (750, 280), "CORE_PCE_YOY": (750, 380),
        "UNEMPLOYMENT": (120, 400), "INITIAL_CLAIMS": (120, 300),
        "NFP": (260, 350), "ISM_MFG": (260, 460),
        "CONSUMER_CONF": (260, 250), "IG_SPREAD": (900, 160),
        "HY_SPREAD": (900, 260),
        "VIX": (750, 120), "SP500": (550, 240), "DOW": (550, 380),
        "NASDAQ": (550, 480), "FTSE100": (1000, 420), "NIKKEI": (1000, 320),
        "OIL_WTI": (420, 420), "OIL_BRT": (420, 520),
        "GOLD": (900, 460), "COPPER": (750, 480),
        "DXY": (900, 360), "PUT_CALL": (750, 180),
    }
    nodes = []
    for key, ind in indicators.items():
        sig = ind.get("signal", "NEUTRAL")
        ext = compute_extremity(ind)
        radius = 15 + ext * 15
        ix, iy = positions.get(key, (550, 300))
        nodes.append({
            "id": key,
            "label": ind.get("label", key),
            "short": short_labels.get(key, key[:6]),
            "signal": sig,
            "color": sig_colors.get(sig, "#ffd24d"),
            "radius": round(radius, 1),
            "ix": ix, "iy": iy,
        })
    edges = [
        {"source": "FED_FUNDS", "target": "US_2Y", "label": "rate transmission"},
        {"source": "US_2Y", "target": "US_10Y", "label": "rate expectations"},
        {"source": "US_10Y", "target": "SPREAD_2S10S", "label": "curve shape"},
        {"source": "OIL_WTI", "target": "CPI_YOY", "label": "energy → inflation"},
        {"source": "CPI_YOY", "target": "CORE_PCE_YOY", "label": "headline → core"},
        {"source": "CORE_PCE_YOY", "target": "FED_FUNDS", "label": "inflation → policy"},
        {"source": "OIL_WTI", "target": "SP500", "label": "margin pressure"},
        {"source": "VIX", "target": "SP500", "label": "fear → equities"},
        {"source": "SP500", "target": "VIX", "label": "sell-off → vol spike"},
        {"source": "UNEMPLOYMENT", "target": "CONSUMER_CONF", "label": "jobs → sentiment"},
        {"source": "HY_SPREAD", "target": "VIX", "label": "credit stress → vol"},
        {"source": "DXY", "target": "COPPER", "label": "dollar → commodities"},
        {"source": "DXY", "target": "GOLD", "label": "dollar → gold"},
        {"source": "GOLD", "target": "US_10Y", "label": "real yields link"},
        {"source": "OIL_BRT", "target": "OIL_WTI", "label": "global → US crude"},
        {"source": "NFP", "target": "UNEMPLOYMENT", "label": "payrolls → jobs"},
        {"source": "ISM_MFG", "target": "SP500", "label": "PMI → equities"},
        {"source": "US_10Y", "target": "SP500", "label": "yields → valuations"},
        {"source": "IG_SPREAD", "target": "HY_SPREAD", "label": "credit contagion"},
    ]
    node_ids = {n["id"] for n in nodes}
    edges = [e for e in edges if e["source"] in node_ids and e["target"] in node_ids]
    return json.dumps(nodes, separators=(',', ':')), json.dumps(edges, separators=(',', ':'))


# ─────────────────────────────────────────────────────────────────────────────
# 6c. MACRO REGIME CLASSIFICATION
# ─────────────────────────────────────────────────────────────────────────────

REGIME_COLORS = {
    "EXPANSION": "#00e676", "TIGHTENING": "#ff4d4d", "SLOWDOWN": "#ffd24d",
    "RECESSION": "#ff1a1a", "RECOVERY": "#4d9fff", "STAGFLATION": "#ff6b35",
    "MIXED": "#8892b0",
}


def classify_regime_from_values(ff=None, ff_prev=None, unemp=None, unemp_prev=None,
                                cpi=None, ism=None, sp=None, sp_prev=None, oil=None,
                                sp_sma200=None):
    """Classify macro regime from indicator values.

    Priority order (first match wins):
    1. RECESSION   — PMI < 47 AND unemployment rising fast
    2. STAGFLATION — oil > $85 AND inflation > 2.5% AND unemployment rising
    3. TIGHTENING  — Fed rate rising AND inflation > 3%
    4. EXPANSION   — S&P above 200-SMA AND unemployment < 4.2% AND (PMI > 50 or None)
    5. RECOVERY    — Fed cutting AND equities rising AND unemployment stable/falling
    6. SLOWDOWN    — (PMI falling toward 50 or consumer weak) AND equities flat/down
    7. MIXED       — fallback
    """
    ff_rising = ff is not None and ff_prev is not None and ff > ff_prev + 0.01
    ff_falling = ff is not None and ff_prev is not None and ff < ff_prev - 0.01
    unemp_rising = (unemp is not None and unemp_prev is not None
                    and unemp - unemp_prev >= 0.1)
    unemp_stable_or_falling = (unemp is not None and unemp_prev is not None
                               and unemp <= unemp_prev + 0.05)
    sp_rising = sp is not None and sp_prev is not None and sp > sp_prev
    sp_above_sma200 = (sp is not None and sp_sma200 is not None
                       and sp > sp_sma200)

    # 1. RECESSION — clear contraction
    if ism is not None and ism < 47 and unemp_rising:
        return "RECESSION"

    # 2. STAGFLATION — oil shock + sticky inflation + labour weakness
    if (oil is not None and oil > 85 and cpi is not None and cpi > 2.5
            and (unemp_rising or (ism is not None and ism < 50))):
        return "STAGFLATION"
    # Also catch: oil > 85 AND inflation > 2.5% even without labour data
    if oil is not None and oil > 85 and cpi is not None and cpi > 2.5:
        return "STAGFLATION"

    # 3. TIGHTENING — Fed hiking into inflation
    if ff_rising and cpi is not None and cpi > 3.0:
        return "TIGHTENING"
    # Broader: Fed rate elevated AND rising even if inflation moderate
    if ff_rising and cpi is not None and cpi > 2.5:
        return "TIGHTENING"

    # 4. EXPANSION — broad strength
    if (sp_above_sma200 and unemp is not None and unemp < 4.2
            and (ism is None or ism > 50)):
        return "EXPANSION"
    # Fallback expansion: strong equities + low unemployment
    if sp_rising and unemp is not None and unemp < 4.0:
        return "EXPANSION"

    # 5. RECOVERY — easing cycle lifting markets
    if ff_falling and sp_rising and unemp_stable_or_falling:
        return "RECOVERY"
    # Broader: equities rising + unemployment not worsening
    if sp_rising and unemp_stable_or_falling and not ff_rising:
        return "RECOVERY"

    # 6. SLOWDOWN — PMI weak, equities struggling
    if ism is not None and ism < 50 and not sp_rising:
        return "SLOWDOWN"
    if ism is not None and ism < 52 and not sp_rising:
        return "SLOWDOWN"

    return "MIXED"


def build_regime_timeline_data(indicators):
    """Build monthly regime timeline for last 24 months."""
    hist_map = {}
    for key in ("FED_FUNDS", "UNEMPLOYMENT", "CPI_YOY", "ISM_MFG", "SP500", "OIL_WTI"):
        hist_map[key] = {}
        for d, v in indicators.get(key, {}).get("history", []):
            hist_map[key][d[:7]] = v
    # Build a rough SMA200 proxy from SP500 history (use 200-day trailing avg)
    sp_history = indicators.get("SP500", {}).get("history", [])
    sp_sma200_map = {}
    if len(sp_history) >= 200:
        from collections import defaultdict
        # Monthly average of the last 200 data points up to that month
        for i in range(200, len(sp_history)):
            ym = sp_history[i][0][:7]
            window = [p[1] for p in sp_history[max(0, i - 199):i + 1]]
            sp_sma200_map[ym] = sum(window) / len(window)
    sp_sma200_current = indicators.get("SP500", {}).get("sma200")

    today = datetime.now(timezone.utc).date()
    months = []
    for i in range(23, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        months.append(f"{y}-{m:02d}")
    timeline = []
    prev_vals = {k: None for k in hist_map}
    for ym in months:
        vals = {k: hist_map[k].get(ym) for k in hist_map}
        sma200 = sp_sma200_map.get(ym, sp_sma200_current)
        regime = classify_regime_from_values(
            ff=vals.get("FED_FUNDS"), ff_prev=prev_vals.get("FED_FUNDS"),
            unemp=vals.get("UNEMPLOYMENT"), unemp_prev=prev_vals.get("UNEMPLOYMENT"),
            cpi=vals.get("CPI_YOY"), ism=vals.get("ISM_MFG"),
            sp=vals.get("SP500"), sp_prev=prev_vals.get("SP500"),
            oil=vals.get("OIL_WTI"), sp_sma200=sma200,
        )
        timeline.append({
            "month": ym,
            "regime": regime,
            "color": REGIME_COLORS.get(regime, "#8892b0"),
        })
        for k in hist_map:
            if vals.get(k) is not None:
                prev_vals[k] = vals[k]
    return timeline


def build_current_regime_description(indicators):
    """Generate current regime label and description."""
    cpi = indicators.get("CPI_YOY", {}).get("value")
    pce = indicators.get("CORE_PCE_YOY", {}).get("value")
    unemp = indicators.get("UNEMPLOYMENT", {}).get("value")
    unemp_prev = indicators.get("UNEMPLOYMENT", {}).get("previous")
    ism = indicators.get("ISM_MFG", {}).get("value")
    ff = indicators.get("FED_FUNDS", {}).get("value")
    ff_prev = indicators.get("FED_FUNDS", {}).get("previous")
    sp = indicators.get("SP500", {}).get("value")
    sp_prev = indicators.get("SP500", {}).get("previous")
    oil = indicators.get("OIL_WTI", {}).get("value")

    sp_sma200 = indicators.get("SP500", {}).get("sma200")
    regime = classify_regime_from_values(
        ff=ff, ff_prev=ff_prev, unemp=unemp, unemp_prev=unemp_prev,
        cpi=cpi, ism=ism, sp=sp, sp_prev=sp_prev, oil=oil,
        sp_sma200=sp_sma200,
    )
    color = REGIME_COLORS.get(regime, "#8892b0")
    parts = [f"Current regime: <strong style='color:{color}'>{regime}</strong>"]
    if regime == "STAGFLATION":
        desc = "Inflation elevated"
        if cpi:
            desc += f" (CPI {cpi:.1f}%)"
        if oil:
            desc += f" with oil at ${oil:.0f}"
        if ism:
            desc += f" while manufacturing weakens (ISM {ism:.1f})"
        parts.append(desc.strip() + ". Last similar period: 2022 Q2-Q3.")
    elif regime == "TIGHTENING":
        desc = "Fed in restrictive mode"
        if ff:
            desc += f" at {ff:.2f}%"
        if cpi:
            desc += f", inflation at {cpi:.1f}%"
        parts.append(desc + ". Bond market pricing sustained higher rates.")
    elif regime == "EXPANSION":
        desc = "Broad growth"
        if unemp:
            desc += f" with unemployment at {unemp:.1f}%"
        if ism:
            desc += f", ISM at {ism:.1f}"
        parts.append(desc + ". Risk assets favored.")
    elif regime == "SLOWDOWN":
        desc = "Growth decelerating"
        if ism:
            desc += f" — ISM at {ism:.1f}"
        if unemp:
            desc += f", unemployment at {unemp:.1f}%"
        parts.append(desc + ". Watch for recession signals.")
    elif regime == "RECESSION":
        desc = "Recessionary conditions"
        if unemp:
            desc += f" — unemployment at {unemp:.1f}%"
        parts.append(desc + ". Defensive positioning warranted.")
    elif regime == "RECOVERY":
        parts.append("Early recovery phase. Equities typically lead the upturn.")
    else:
        parts.append("Mixed signals — no dominant regime. Multiple crosscurrents.")
    return regime, color, " — ".join(parts)


# ─────────────────────────────────────────────────────────────────────────────
# 6d. SCENARIO BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_scenarios(indicators):
    """Return list of scenario dicts based on current data."""
    scenarios = []
    oil = indicators.get("OIL_WTI", {}).get("value")
    vix = indicators.get("VIX", {}).get("value")
    sp = indicators.get("SP500", {}).get("value")
    pce = indicators.get("CORE_PCE_YOY", {}).get("value")
    cpi = indicators.get("CPI_YOY", {}).get("value")
    ff = indicators.get("FED_FUNDS", {}).get("value")
    ten_y = indicators.get("US_10Y", {}).get("value")
    dxy = indicators.get("DXY", {}).get("value")
    gold = indicators.get("GOLD", {}).get("value")

    if oil and oil > 90:
        sp_lo = f"{sp * 0.93:,.0f}" if sp else "\u2014"
        sp_hi = f"{sp * 0.96:,.0f}" if sp else "\u2014"
        cpi_lo = f"{cpi + 0.3:.1f}" if cpi else "?"
        cpi_hi = f"{cpi + 0.5:.1f}" if cpi else "?"
        ff_str = f"{ff:.2f}%" if ff else "current level"
        scenarios.append({
            "title": "Oil Stays Above $100",
            "probability": "Medium" if oil > 95 else "Low",
            "prob_color": "#ffd24d" if oil > 95 else "#8892b0",
            "border": "#ff4d4d",
            "content": (f"If WTI (currently ${oil:.0f}) remains above $100 for 4+ weeks: "
                        f"CPI likely rises to {cpi_lo}\u2013{cpi_hi}%, "
                        f"consumer spending contracts 1\u20132%, "
                        f"Fed forced to hold at {ff_str} or tighten, "
                        f"S&amp;P 500 downside {sp_lo}\u2013{sp_hi} range. "
                        f"Historical precedent: 2022 oil shock saw 3-month lag "
                        f"to peak CPI impact."),
            "impacts": [("CPI", "\u25b2", "#ff4d4d"), ("S&amp;P 500", "\u25bc", "#ff4d4d"),
                        ("Fed Funds", "\u25b2", "#ff4d4d"), ("Gold", "\u25b2", "#00e676")],
        })

    if ff and ff > 4.0:
        ten_lo = f"{ten_y - 0.3:.2f}" if ten_y else "?"
        ten_hi = f"{ten_y - 0.2:.2f}" if ten_y else "?"
        sp_lo2 = f"{sp * 1.03:,.0f}" if sp else "?"
        sp_hi2 = f"{sp * 1.05:,.0f}" if sp else "?"
        dxy_lo = f"{dxy - 3:.0f}" if dxy else "?"
        dxy_hi = f"{dxy - 2:.0f}" if dxy else "?"
        prob = "Low" if (pce and pce > 2.5) else "Medium"
        pce_str = f"{pce:.1f}%" if pce else "elevated"
        scenarios.append({
            "title": "Fed Cuts by June",
            "probability": prob,
            "prob_color": "#8892b0" if prob == "Low" else "#ffd24d",
            "border": "#00e676",
            "content": (f"If Fed delivers surprise cut from {ff:.2f}%: "
                        f"10Y likely drops to {ten_lo}\u2013{ten_hi}%, "
                        f"S&amp;P rallies to {sp_lo2}\u2013{sp_hi2} in first week, "
                        f"DXY weakens toward {dxy_lo}\u2013{dxy_hi}, gold benefits. "
                        f"Probability currently {prob.lower()} given Core PCE at {pce_str}."),
            "impacts": [("10Y Yield", "\u25bc", "#00e676"), ("S&amp;P 500", "\u25b2", "#00e676"),
                        ("DXY", "\u25bc", "#00e676"), ("Gold", "\u25b2", "#00e676")],
        })

    if oil and oil > 85:
        vix_str = f"{vix:.1f}" if vix else "?"
        sp_lo3 = f"{sp * 1.05:,.0f}" if sp else "?"
        sp_hi3 = f"{sp * 1.08:,.0f}" if sp else "?"
        scenarios.append({
            "title": "Geopolitical De-escalation",
            "probability": "Low",
            "prob_color": "#8892b0",
            "border": "#00e676",
            "content": (f"If geopolitical tensions ease: oil (currently ${oil:.0f}) "
                        f"could drop to $70\u201375, VIX (currently {vix_str}) back below 18, "
                        f"equities rally to {sp_lo3}\u2013{sp_hi3} on multiple expansion, "
                        f"credit spreads tighten 50\u201380bps."),
            "impacts": [("Oil", "\u25bc", "#00e676"), ("VIX", "\u25bc", "#00e676"),
                        ("S&amp;P 500", "\u25b2", "#00e676"), ("HY Spread", "\u25bc", "#00e676")],
        })

    return scenarios


# ─────────────────────────────────────────────────────────────────────────────
# 6f. SUMMARY PANELS — Daily Brief, Weekly Wrap, Forward Look
# ─────────────────────────────────────────────────────────────────────────────

ECONOMIC_CALENDAR = [
    {"date": "2026-03-28", "name": "Core PCE Price Index (Feb)", "category": "Inflation"},
    {"date": "2026-04-03", "name": "Nonfarm Payrolls (Mar)", "category": "Labour"},
    {"date": "2026-04-04", "name": "ISM Services PMI (Mar)", "category": "Activity"},
    {"date": "2026-04-10", "name": "CPI (Mar)", "category": "Inflation"},
    {"date": "2026-04-16", "name": "Retail Sales (Mar)", "category": "Consumption"},
    {"date": "2026-04-30", "name": "GDP Advance Q1", "category": "Growth"},
    {"date": "2026-05-01", "name": "ISM Manufacturing PMI (Apr)", "category": "Activity"},
    {"date": "2026-05-02", "name": "Nonfarm Payrolls (Apr)", "category": "Labour"},
    {"date": "2026-05-05", "name": "FOMC Meeting Day 1", "category": "Rates"},
    {"date": "2026-05-06", "name": "FOMC Decision", "category": "Rates"},
    {"date": "2026-05-13", "name": "CPI (Apr)", "category": "Inflation"},
]


def _biggest_mover(indicators):
    """Return (key, label, pct_change, direction, value_str) of biggest % mover."""
    best_key, best_pct = None, 0
    relative_keys = {"SP500", "DOW", "NASDAQ", "FTSE100", "NIKKEI",
                     "OIL_WTI", "OIL_BRT", "GOLD", "COPPER", "DXY", "VIX"}
    for key, ind in indicators.items():
        v = ind.get("value")
        prev = ind.get("previous")
        if v is None or prev is None or prev == 0:
            continue
        if key in relative_keys:
            pct = abs((v - prev) / prev * 100)
        else:
            pct = abs(v - prev)
        if pct > best_pct:
            best_pct = pct
            best_key = key
    if best_key is None:
        return None
    ind = indicators[best_key]
    v, prev = ind["value"], ind["previous"]
    diff = v - prev
    adiff = abs(diff)
    unit = ind.get("unit", "")
    if best_key in relative_keys:
        pct = abs(diff / prev * 100)
        chg_str = f"{pct:.1f}%"
    elif best_key == "INITIAL_CLAIMS":
        chg_str = f"{adiff / 1000:.0f}K"
    elif best_key == "NFP":
        chg_str = f"{adiff:.0f}K"
    elif unit == "bps":
        chg_str = f"{adiff:.0f}bps"
    elif unit.endswith("% YoY") or unit == "%":
        chg_str = f"{adiff:.2f}pp"
    else:
        chg_str = f"{adiff:.1f}" if adiff < 100 else f"{adiff:,.0f}"
    direction = "up" if diff > 0 else "down"
    return (best_key, ind.get("label", best_key), chg_str, direction,
            value_display(best_key, ind))


def _calendar_today_tomorrow(today=None):
    """Return (event_name, when_str) for today/tomorrow release, or None."""
    from datetime import date as date_cls
    if today is None:
        today = datetime.now(timezone.utc).date()
    tomorrow = today + timedelta(days=1)
    for evt in ECONOMIC_CALENDAR:
        d = datetime.strptime(evt["date"], "%Y-%m-%d").date()
        if d == today:
            return (evt["name"], "today", evt["category"])
        if d == tomorrow:
            return (evt["name"], "tomorrow", evt["category"])
    return None


def _calendar_next_week(today=None):
    """Return list of (name, date_str) for releases in the next 7 days."""
    from datetime import date as date_cls
    if today is None:
        today = datetime.now(timezone.utc).date()
    start = today + timedelta(days=1)
    end = today + timedelta(days=8)
    result = []
    for evt in ECONOMIC_CALENDAR:
        d = datetime.strptime(evt["date"], "%Y-%m-%d").date()
        if start <= d <= end:
            result.append((evt["name"], d.strftime("%-d %b"), evt["category"]))
    return result[:3]


def generate_daily_brief(indicators):
    """Generate 4-sentence daily brief from current indicator data."""
    bull = sum(1 for i in indicators.values() if i.get("signal") == "BULLISH")
    bear = sum(1 for i in indicators.values() if i.get("signal") == "BEARISH")

    # Sentence 1 — Regime statement
    if bear > bull + 5:
        s1 = f"Risk-off dominates with {bear} bearish signals."
    elif bull > bear + 5:
        s1 = f"Risk-on sentiment with {bull} bullish signals."
    else:
        s1 = f"Mixed signals across the board — {bull} bullish vs {bear} bearish."

    # Sentence 2 — Biggest driver
    mover = _biggest_mover(indicators)
    if mover:
        mk, mlabel, mchg, mdir, mval = mover
        commentary_short = indicators[mk].get("commentary", "")
        # Take first sentence of commentary
        first_sent = commentary_short.split(". ")[0] + "." if ". " in commentary_short else commentary_short
        s2 = f"Biggest mover: {mlabel}, {mdir} {mchg} to {mval}. {first_sent}"
    else:
        s2 = "No significant moves since last refresh."

    # Sentence 3 — Cross-indicator read
    oil = indicators.get("OIL_WTI", {}).get("value")
    cpi = indicators.get("CPI_YOY", {}).get("value")
    vix = indicators.get("VIX", {}).get("value")
    hy = indicators.get("HY_SPREAD", {}).get("value")
    unemp = indicators.get("UNEMPLOYMENT", {}).get("value")
    unemp_prev = indicators.get("UNEMPLOYMENT", {}).get("previous")
    cons = indicators.get("CONSUMER_CONF", {}).get("value")
    cons_prev = indicators.get("CONSUMER_CONF", {}).get("previous")
    ten_y = indicators.get("US_10Y", {}).get("value")
    sp = indicators.get("SP500", {}).get("value")
    sp_prev = indicators.get("SP500", {}).get("previous")

    s3 = None
    if oil and oil > 90 and cpi and cpi > 2.5:
        s3 = (f"Oil above ${oil:.0f} is feeding inflation expectations "
              f"— stagflation risk elevated.")
    elif vix and vix > 25 and hy and hy > 400:
        s3 = (f"Elevated VIX ({vix:.1f}) and widening high-yield spreads "
              f"({hy:.0f}bps) signal institutional risk aversion.")
    elif (unemp and unemp_prev and unemp > unemp_prev
          and cons and cons_prev and cons < cons_prev):
        s3 = ("Labour market softening while consumer confidence drops "
              "— demand weakness building.")
    elif (ten_y and sp and sp_prev and sp < sp_prev
          and indicators.get("US_10Y", {}).get("direction") == "▲"):
        s3 = (f"Rising yields ({ten_y:.2f}%) pressuring equities "
              "— growth/tech most vulnerable.")
    elif (oil and oil < 75 and sp and sp_prev and sp > sp_prev):
        s3 = ("Falling oil providing relief to margins "
              "— supportive backdrop for equities.")
    if not s3:
        s3 = ("No dominant cross-indicator theme today "
              "— monitor individual signals.")

    # Sentence 4 — What to watch
    cal = _calendar_today_tomorrow()
    if cal:
        s4 = f"Key catalyst: {cal[0]} due {cal[1]}."
    else:
        s4 = ("No major releases today. Watch oil prices and credit "
              "spreads for sentiment shifts.")

    return f"{s1} {s2} {s3} {s4}"


def generate_weekly_wrap(indicators, previous_data):
    """Generate 5-sentence weekly wrap. Call only on Fri/Sat/Sun."""
    # Compare current to ~1 week ago using history
    improved, deteriorated, unchanged = 0, 0, 0
    movers = []
    for key, ind in indicators.items():
        h = ind.get("history", [])
        v = ind.get("value")
        if v is None or len(h) < 6:
            unchanged += 1
            continue
        week_ago_val = h[-6][1] if len(h) >= 6 else h[0][1]
        if week_ago_val == 0:
            unchanged += 1
            continue
        diff = v - week_ago_val
        pct = abs(diff / week_ago_val * 100) if week_ago_val != 0 else 0
        sig = ind.get("signal", "NEUTRAL")
        # Improved = moved in bullish direction
        if sig == "BULLISH" and diff > 0:
            improved += 1
        elif sig == "BEARISH" and diff < 0:
            improved += 1
        elif abs(diff / week_ago_val * 100) < 0.2 if week_ago_val != 0 else True:
            unchanged += 1
        else:
            deteriorated += 1
        movers.append((key, ind.get("label", key), pct, diff, v,
                        value_display(key, ind)))

    movers.sort(key=lambda x: x[2], reverse=True)

    # Sentence 1 — Scorecard
    s1 = (f"This week: {improved} indicators improved, "
          f"{deteriorated} deteriorated, {unchanged} unchanged.")

    # Sentence 2 — Top 2 movers
    # For NFP and similar indicators, use absolute change (% is meaningless
    # when values cross zero).
    abs_change_keys = {"NFP", "INITIAL_CLAIMS", "ISM_MFG", "CONSUMER_CONF"}
    inflation_keys = {"CPI_YOY", "CORE_CPI_YOY", "CORE_PCE_YOY"}
    spread_keys = {"IG_SPREAD", "HY_SPREAD"}

    def _mover_str(m):
        key, label, pct, diff, v, val_str = m
        d = "up" if diff > 0 else "down"
        adiff = abs(diff)
        if key in inflation_keys:
            return f"{label} moved most, {d} {adiff:.2f}pp to {val_str}"
        if key in spread_keys:
            return f"{label} moved most, {d} {adiff:.0f}bps to {val_str}"
        if key in abs_change_keys:
            unit = indicators.get(key, {}).get("unit", "")
            return f"{label} moved most, {d} {adiff:,.0f}{unit} to {val_str}"
        return f"{label} moved most, {d} {pct:.1f}% to {val_str}"

    if len(movers) >= 2:
        m1, m2 = movers[0], movers[1]
        s2 = f"{_mover_str(m1)}. {_mover_str(m2).replace('moved most', 'also notable')}."
    elif len(movers) == 1:
        s2 = f"{_mover_str(movers[0])}."
    else:
        s2 = "No significant movers this week."

    # Sentence 3 — Regime shift check
    bull = sum(1 for i in indicators.values() if i.get("signal") == "BULLISH")
    bear = sum(1 for i in indicators.values() if i.get("signal") == "BEARISH")
    prev_bull = sum(1 for i in previous_data.values()
                    if isinstance(i, dict) and i.get("signal") == "BULLISH")
    prev_bear = sum(1 for i in previous_data.values()
                    if isinstance(i, dict) and i.get("signal") == "BEARISH")
    bear_delta = bear - prev_bear
    bull_delta = bull - prev_bull
    if bear_delta >= 3:
        s3 = (f"Macro backdrop deteriorated this week — "
              f"{bear_delta} indicators flipped bearish.")
    elif bull_delta >= 3:
        s3 = (f"Conditions improved — "
              f"{bull_delta} indicators turned bullish.")
    else:
        s3 = "Overall macro regime unchanged week-over-week."

    # Sentence 4 — Key theme
    oil = indicators.get("OIL_WTI", {}).get("value")
    cpi = indicators.get("CPI_YOY", {}).get("value")
    vix = indicators.get("VIX", {}).get("value")
    hy = indicators.get("HY_SPREAD", {}).get("value")
    if oil and oil > 90 and cpi and cpi > 2.5:
        s4 = "The week was defined by persistent oil-driven inflation pressure."
    elif vix and vix > 25:
        s4 = f"The week was defined by elevated volatility (VIX {vix:.1f})."
    elif bear > bull + 3:
        s4 = "The week was defined by broadening bearish sentiment."
    elif bull > bear + 3:
        s4 = "The week was defined by improving risk appetite."
    else:
        s4 = "No single theme dominated — crosscurrents persisted."

    # Sentence 5 — Next week preview
    next_events = _calendar_next_week()
    if next_events:
        previews = [f"{n} ({d})" for n, d, _ in next_events[:2]]
        cats = set(c for _, _, c in next_events[:2])
        cat_str = "/".join(sorted(cats)).lower()
        s5 = (f"Next week: {', '.join(previews)}. "
              f"These could shift the {cat_str} picture.")
    else:
        s5 = "Light calendar next week — watch for policy commentary."

    return f"{s1} {s2} {s3} {s4} {s5}"


def generate_forward_look(indicators):
    """Generate conditional scenario trees. Returns list of scenario dicts."""
    oil = indicators.get("OIL_WTI", {}).get("value")
    vix = indicators.get("VIX", {}).get("value")
    sp = indicators.get("SP500", {}).get("value")
    cpi = indicators.get("CPI_YOY", {}).get("value")
    core_cpi = indicators.get("CORE_CPI_YOY", {}).get("value")
    pce = indicators.get("CORE_PCE_YOY", {}).get("value")
    nfp = indicators.get("NFP", {}).get("value")
    unemp = indicators.get("UNEMPLOYMENT", {}).get("value")
    ten_y = indicators.get("US_10Y", {}).get("value")
    dxy = indicators.get("DXY", {}).get("value")
    hy = indicators.get("HY_SPREAD", {}).get("value")
    gold = indicators.get("GOLD", {}).get("value")
    ff = indicators.get("FED_FUNDS", {}).get("value")

    scenarios = []

    # Oil shock persistence
    if oil and oil > 85:
        sp_support = f"{sp * 0.95:,.0f}" if sp else "?"
        wti_r = f"{oil:.0f}"
        scenarios.append({
            "title": "Oil shock persistence",
            "probability": "HIGH" if oil > 95 else "MEDIUM",
            "prob_color": "#ff4d4d" if oil > 95 else "#ffd24d",
            "text": (f"If oil holds above ${wti_r} through month-end, expect next CPI "
                     f"to print 2.6\u20132.8% \u2014 above consensus. This would push "
                     f"rate cut expectations to zero for H1 2026 and pressure equities "
                     f"toward {sp_support} support."),
            "impacts": [("S&P", "\u25bc", "#ff4d4d"), ("CPI", "\u25b2", "#ff4d4d"),
                        ("10Y", "\u25b2", "#ff4d4d"), ("Gold", "\u25b2", "#00e676"),
                        ("VIX", "\u25b2", "#ff4d4d")],
        })

    # Labour market cracks
    if (nfp is not None and nfp < 50) or (unemp and unemp > 4.3):
        ten_target = f"{ten_y - 0.3:.2f}" if ten_y else "?"
        sp_test = f"{sp * 0.92:,.0f}" if sp else "?"
        scenarios.append({
            "title": "Labour market cracks",
            "probability": "MEDIUM" if (nfp is not None and nfp < 0) else "LOW",
            "prob_color": "#ffd24d" if (nfp is not None and nfp < 0) else "#8892b0",
            "text": (f"If next NFP prints below zero again, recession odds jump "
                     f"to ~50%. Expect flight to bonds (10Y toward {ten_target}%), "
                     f"gold rally, and S&P testing {sp_test}."),
            "impacts": [("S&P", "\u25bc", "#ff4d4d"), ("10Y", "\u25bc", "#00e676"),
                        ("Gold", "\u25b2", "#00e676"), ("VIX", "\u25b2", "#ff4d4d"),
                        ("DXY", "\u25bc", "#00e676")],
        })

    # Inflation re-acceleration
    if cpi and cpi > 2.3 and oil and oil > 80:
        ten_up = f"{ten_y + 0.3:.2f}" if ten_y else "?"
        cc = f"{core_cpi:.1f}" if core_cpi else "?"
        scenarios.append({
            "title": "Inflation re-acceleration",
            "probability": "MEDIUM" if cpi > 2.8 else "LOW",
            "prob_color": "#ffd24d" if cpi > 2.8 else "#8892b0",
            "text": (f"If shelter costs stay sticky ({cc}%) and oil adds 0.3\u20130.5pp "
                     f"to headline CPI, expect Fed rhetoric to turn more hawkish. "
                     f"10Y could test {ten_up}% and growth stocks underperform value."),
            "impacts": [("10Y", "\u25b2", "#ff4d4d"), ("S&P", "\u25bc", "#ff4d4d"),
                        ("Gold", "\u25b2", "#00e676"), ("DXY", "\u25b2", "#ff4d4d"),
                        ("CPI", "\u25b2", "#ff4d4d")],
        })

    # Geopolitical de-escalation
    if oil and oil > 85:
        scenarios.append({
            "title": "Geopolitical de-escalation",
            "probability": "LOW",
            "prob_color": "#8892b0",
            "text": (f"If Iran conflict resolves and Strait reopens, oil could drop "
                     f"to $70\u201375. VIX back below 18, equities rally 5\u20138%. "
                     f"Energy stocks give back gains, consumer discretionary outperforms."),
            "impacts": [("Oil", "\u25bc", "#00e676"), ("VIX", "\u25bc", "#00e676"),
                        ("S&P", "\u25b2", "#00e676"), ("Gold", "\u25bc", "#ff4d4d"),
                        ("DXY", "\u25bc", "#00e676")],
        })

    # Soft landing confirmed
    if unemp and unemp < 4.5 and cpi and cpi < 2.5:
        sp_target = f"{sp * 1.08:,.0f}" if sp else "?"
        scenarios.append({
            "title": "Soft landing confirmed",
            "probability": "LOW" if (cpi and cpi > 2.3) else "MEDIUM",
            "prob_color": "#8892b0" if (cpi and cpi > 2.3) else "#ffd24d",
            "text": (f"Unemployment stable at {unemp:.1f}% with inflation near target "
                     f"\u2014 textbook soft landing. If this holds, S&P targets "
                     f"{sp_target} by year-end with gradual rate normalisation."),
            "impacts": [("S&P", "\u25b2", "#00e676"), ("10Y", "\u25bc", "#00e676"),
                        ("VIX", "\u25bc", "#00e676"), ("Gold", "\u25ac", "#ffd24d"),
                        ("DXY", "\u25ac", "#ffd24d")],
        })

    # Credit stress escalation (hy is now in bps)
    if hy and hy > 400:
        scenarios.append({
            "title": "Credit stress escalation",
            "probability": "MEDIUM" if hy > 450 else "LOW",
            "prob_color": "#ffd24d" if hy > 450 else "#8892b0",
            "text": (f"HY spreads at {hy:.0f}bps and rising. If they breach 500bps, "
                     f"corporate borrowing freezes up \u2014 historically precedes "
                     f"15\u201320% equity drawdown within 3 months. "
                     f"Watch CCC-rated issuers for first defaults."),
            "impacts": [("S&P", "\u25bc", "#ff4d4d"), ("HY", "\u25b2", "#ff4d4d"),
                        ("VIX", "\u25b2", "#ff4d4d"), ("Gold", "\u25b2", "#00e676"),
                        ("10Y", "\u25bc", "#00e676")],
        })

    # Dollar breakdown
    if dxy and dxy < 100:
        scenarios.append({
            "title": "Dollar breakdown",
            "probability": "MEDIUM" if dxy < 98 else "LOW",
            "prob_color": "#ffd24d" if dxy < 98 else "#8892b0",
            "text": (f"DXY below 100 at {dxy:.1f}. Further weakness toward 96\u201397 "
                     f"would boost EM equities and commodities but signals fading "
                     f"confidence in US growth exceptionalism."),
            "impacts": [("DXY", "\u25bc", "#ff4d4d"), ("Gold", "\u25b2", "#00e676"),
                        ("Copper", "\u25b2", "#00e676"), ("Oil", "\u25b2", "#ffd24d"),
                        ("S&P", "\u25ac", "#ffd24d")],
        })

    # Fed surprise cut — always-available tail risk (low priority)
    if ff and ten_y:
        scenarios.append({
            "title": "Fed surprise cut",
            "probability": "LOW",
            "prob_color": "#8892b0",
            "text": (f"A surprise 25bp cut (currently <10% probability) would send "
                     f"10Y to ~{ten_y - 0.25:.2f}%, S&P rallying 3\u20135% in a week, "
                     f"DXY below 98. Trigger: sharp NFP miss or credit event."),
            "impacts": [("S&P", "\u25b2", "#00e676"), ("10Y", "\u25bc", "#00e676"),
                        ("DXY", "\u25bc", "#00e676"), ("Gold", "\u25b2", "#00e676"),
                        ("VIX", "\u25bc", "#00e676")],
        })

    return scenarios[:3]  # First 3 matching


# ─────────────────────────────────────────────────────────────────────────────
# 6e. D3 MACRO MAP JS (regular string — no f-string escaping needed)
# ─────────────────────────────────────────────────────────────────────────────

D3_MACRO_MAP_JS = """
function renderMacroMap() {
    const container = document.getElementById('macro-map-container');
    if (!container) return;
    const width = container.clientWidth || 1200;
    const height = 600;
    container.innerHTML = '';

    const svg = d3.select(container).append('svg')
        .attr('width', width).attr('height', height);

    // Arrow marker
    svg.append('defs').append('marker')
        .attr('id', 'arrowhead').attr('viewBox', '0 -5 10 10')
        .attr('refX', 25).attr('refY', 0)
        .attr('markerWidth', 6).attr('markerHeight', 6)
        .attr('orient', 'auto')
        .append('path').attr('d', 'M0,-5L10,0L0,5').attr('fill', '#667194');

    const nodes = JSON.parse(JSON.stringify(MACRO_MAP_NODES));
    const edges = JSON.parse(JSON.stringify(MACRO_MAP_EDGES));

    const simulation = d3.forceSimulation(nodes)
        .force('link', d3.forceLink(edges).id(d => d.id).distance(130).strength(0.4))
        .force('charge', d3.forceManyBody().strength(-350))
        .force('center', d3.forceCenter(width / 2, height / 2))
        .force('collision', d3.forceCollide().radius(d => d.radius + 8))
        .force('x', d3.forceX(d => d.ix * width / 1100).strength(0.15))
        .force('y', d3.forceY(d => d.iy * height / 600).strength(0.15));

    const link = svg.selectAll('.mm-link-g').data(edges).enter().append('g');
    const linkLine = link.append('line')
        .attr('stroke', '#667194').attr('stroke-width', 1.5)
        .attr('stroke-opacity', 0.35).attr('marker-end', 'url(#arrowhead)');
    const linkLabel = link.append('text')
        .attr('text-anchor', 'middle').attr('fill', '#4a5278')
        .attr('font-size', '8px').attr('font-family', "'DM Sans', sans-serif")
        .text(d => d.label);

    const node = svg.selectAll('.mm-node').data(nodes).enter().append('g')
        .style('cursor', 'pointer')
        .call(d3.drag()
            .on('start', (e, d) => { if (!e.active) simulation.alphaTarget(0.3).restart(); d.fx = d.x; d.fy = d.y; })
            .on('drag', (e, d) => { d.fx = e.x; d.fy = e.y; })
            .on('end', (e, d) => { if (!e.active) simulation.alphaTarget(0); d.fx = null; d.fy = null; }));

    node.append('circle')
        .attr('r', d => d.radius)
        .attr('fill', d => d.color + '33')
        .attr('stroke', d => d.color).attr('stroke-width', 2);
    node.append('text')
        .attr('text-anchor', 'middle').attr('dy', '0.35em')
        .attr('fill', '#e8eaf2')
        .attr('font-size', d => d.radius > 22 ? '10px' : '8px')
        .attr('font-family', "'JetBrains Mono', monospace")
        .attr('font-weight', '600').text(d => d.short);

    node.on('mouseover', function(event, d) {
        const connected = new Set([d.id]);
        edges.forEach(e => {
            const s = typeof e.source === 'object' ? e.source.id : e.source;
            const t = typeof e.target === 'object' ? e.target.id : e.target;
            if (s === d.id) connected.add(t);
            if (t === d.id) connected.add(s);
        });
        node.select('circle').attr('opacity', n => connected.has(n.id) ? 1 : 0.12);
        node.select('text').attr('opacity', n => connected.has(n.id) ? 1 : 0.12);
        linkLine.attr('stroke-opacity', e => {
            const s = typeof e.source === 'object' ? e.source.id : e.source;
            const t = typeof e.target === 'object' ? e.target.id : e.target;
            return (s === d.id || t === d.id) ? 0.9 : 0.04;
        }).attr('stroke', e => {
            const s = typeof e.source === 'object' ? e.source.id : e.source;
            const t = typeof e.target === 'object' ? e.target.id : e.target;
            return (s === d.id || t === d.id) ? '#4d9fff' : '#667194';
        });
        linkLabel.attr('opacity', e => {
            const s = typeof e.source === 'object' ? e.source.id : e.source;
            const t = typeof e.target === 'object' ? e.target.id : e.target;
            return (s === d.id || t === d.id) ? 1 : 0.04;
        });
    }).on('mouseout', function() {
        node.select('circle').attr('opacity', 1);
        node.select('text').attr('opacity', 1);
        linkLine.attr('stroke-opacity', 0.35).attr('stroke', '#667194');
        linkLabel.attr('opacity', 1);
    }).on('click', function(event, d) {
        const btn = document.querySelector('.macromap-btn');
        if (btn && btn.classList.contains('active')) toggleMacroMap(btn);
        setTimeout(() => scrollToCard(d.id), 200);
    });

    simulation.on('tick', () => {
        nodes.forEach(d => {
            d.x = Math.max(d.radius + 5, Math.min(width - d.radius - 5, d.x));
            d.y = Math.max(d.radius + 5, Math.min(height - d.radius - 5, d.y));
        });
        linkLine.attr('x1', d => d.source.x).attr('y1', d => d.source.y)
            .attr('x2', d => d.target.x).attr('y2', d => d.target.y);
        linkLabel.attr('x', d => (d.source.x + d.target.x) / 2)
            .attr('y', d => (d.source.y + d.target.y) / 2 - 6);
        node.attr('transform', d => 'translate(' + d.x + ',' + d.y + ')');
    });
}

function toggleMacroMap(btn) {
    const mm = document.getElementById('macro-map-section');
    const cv = document.getElementById('cards-view');
    const hm = document.getElementById('heatmap-section');
    const hmBtn = document.querySelector('.heatmap-btn');
    const isActive = btn.classList.contains('active');

    if (!isActive && hmBtn && hmBtn.classList.contains('active')) {
        hmBtn.classList.remove('active');
        hmBtn.textContent = '\\u25a6 Heat Map';
        hm.style.display = 'none';
    }

    btn.classList.toggle('active', !isActive);
    btn.textContent = isActive ? '\\u26A1 Macro Map' : '\\u2715 Cards View';
    mm.style.display = isActive ? 'none' : 'block';
    cv.style.display = isActive ? 'block' : 'none';
    if (hm) hm.style.display = 'none';

    if (!isActive) setTimeout(renderMacroMap, 100);
}
"""


def generate_html(indicators: dict, timestamp: str, *,
                   daily_brief="", weekly_wrap="", forward_look=None) -> str:
    # Build signal counts
    counts = {"BULLISH": 0, "BEARISH": 0, "NEUTRAL": 0}
    for ind in indicators.values():
        s = ind.get("signal", "NEUTRAL")
        counts[s] = counts.get(s, 0) + 1

    # Build history JSON for embedding in HTML
    history_data = {}
    for key, ind in indicators.items():
        h = ind.get("history", [])
        if h:
            history_data[key] = h
    history_json = json.dumps(history_data, separators=(',', ':'))

    # Build narrative
    narrative = generate_narrative(indicators)

    # Build upcoming releases
    upcoming_releases = get_upcoming_releases()

    # Build indicator metadata for JS (labels, directions, values, signals)
    ind_meta = {}
    for key, ind in indicators.items():
        ind_meta[key] = {
            "label": ind.get("label", key),
            "direction": ind.get("direction", "\u25ac"),
            "value": value_display(key, ind),
            "signal": ind.get("signal", "NEUTRAL"),
        }
    ind_meta_json = json.dumps(ind_meta, separators=(',', ':'))
    correlations_json = json.dumps(CORRELATIONS, separators=(',', ':'))

    # Compute heat scores
    heat_colors = {-2: "#ff1a1a", -1: "#ff4d4d", 0: "#3d4663", 1: "#00c853", 2: "#00ff88"}
    heat_data = {}
    for key, ind in indicators.items():
        sig = ind.get("signal", "NEUTRAL")
        base = 1 if sig == "BULLISH" else -1 if sig == "BEARISH" else 0
        momentum = ind.get("momentum")
        if momentum == "ACCEL" and base != 0:
            heat = base + (1 if base > 0 else -1)
        elif momentum == "DECEL" and base != 0:
            heat = 0
        else:
            heat = base
        heat = max(-2, min(2, heat))
        ind["heat_score"] = heat
        heat_data[key] = heat
    heat_json = json.dumps(heat_data, separators=(',', ':'))

    # Detect meaningful changes
    changes_list = detect_changes(indicators)

    # Build "What Changed" as inline sub-section (embedded in Daily Brief)
    if changes_list:
        change_items = ""
        for ck, clabel, cold, cnew, ctx in changes_list:
            change_items += (
                f'<div class="change-item">'
                f'<span class="change-name">{clabel}</span>'
                f'<span class="change-vals">{cold} &rarr; {cnew}</span>'
                f'<span class="change-ctx">{ctx}</span>'
                f'</div>')
        changes_inline_html = (
            f'<div class="changes-inline">'
            f'<div class="changes-inline-header" onclick="toggleChanges()">'
            f'<span>&#9889; Show changes ({len(changes_list)})</span>'
            f'<span class="changes-toggle" id="changes-toggle">&#9654;</span>'
            f'</div>'
            f'<div class="changes-body collapsed" id="changes-body">{change_items}</div>'
            f'</div>')
    else:
        changes_inline_html = ""

    # Build heat map
    heatmap_cells = ""
    for key, ind in indicators.items():
        hs = ind.get("heat_score", 0)
        bg = heat_colors.get(hs, "#3d4663")
        label_short = ind.get("label", key)
        if len(label_short) > 16:
            label_short = label_short[:14] + "&#8230;"
        hm_val = value_display(key, ind)
        hm_sig = ind.get("signal", "NEUTRAL")
        heatmap_cells += (
            f'<div class="hm-cell" data-key="{key}" '
            f'style="background:{bg}" '
            f"onclick=\"event.stopPropagation();switchToCard('{key}')\" "
            f'title="{ind.get("label", key)}: {hm_val} ({hm_sig})">'
            f'<span class="hm-label">{label_short}</span>'
            f'<span class="hm-val">{hm_val}</span>'
            f'</div>')
    heatmap_html = f"""
      <section class="heatmap-section" id="heatmap-section" style="display:none">
        <h2 class="section-title">&#128293; Heat Map</h2>
        <div class="heatmap-grid">{heatmap_cells}</div>
        <div class="hm-legend">
          <span style="color:#ff1a1a">&#9632; Strong Bear</span>
          <span style="color:#ff4d4d">&#9632; Bearish</span>
          <span style="color:#3d4663">&#9632; Neutral</span>
          <span style="color:#00c853">&#9632; Bullish</span>
          <span style="color:#00ff88">&#9632; Strong Bull</span>
        </div>
      </section>"""

    # Build macro map data
    map_nodes_json, map_edges_json = build_macro_map_data(indicators)

    # Build regime timeline
    timeline_data = build_regime_timeline_data(indicators)
    timeline_json = json.dumps(timeline_data, separators=(',', ':'))
    regime_cur, regime_cur_color, regime_desc = build_current_regime_description(indicators)

    # Build scenarios
    scenarios = build_scenarios(indicators)
    scenarios_parts = ""
    for sc in scenarios:
        impacts_items = ""
        for sname, sarrow, scolor in sc["impacts"]:
            impacts_items += f'<span class="sc-impact" style="color:{scolor}">{sarrow} {sname}</span>'
        scenarios_parts += f"""
      <div class="scenario-box" style="border-color:{sc['border']}">
        <div class="sc-header">
          <span class="sc-title">{sc['title']}</span>
          <span class="sc-prob" style="color:{sc['prob_color']}">{sc['probability']}</span>
        </div>
        <p class="sc-content">{sc['content']}</p>
        <div class="sc-impacts">{impacts_items}</div>
      </div>"""
    if scenarios:
        scenarios_html = f"""
      <section class="section scenarios-section">
        <h2 class="section-title">&#128202; Scenarios</h2>
        <div class="scenarios-grid">{scenarios_parts}</div>
      </section>"""
    else:
        scenarios_html = ""

    # Build timeline HTML
    tl_segments = ""
    tl_month_labels = ""
    month_names = {
        "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
        "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
        "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
    }
    for td in timeline_data:
        month_short = td["month"][5:]
        tl_segments += (
            f'<div class="tl-segment" style="background:{td["color"]}" '
            f'title="{td["month"]}: {td["regime"]}">'
            f'<span class="tl-month">{month_short}</span></div>')
        month_num = td["month"][5:7]
        tl_month_labels += f'<span>{month_names.get(month_num, month_num)}</span>'
    used_regimes = list(dict.fromkeys(td["regime"] for td in timeline_data))
    tl_legend = ""
    for r in used_regimes:
        c = REGIME_COLORS.get(r, "#8892b0")
        tl_legend += f'<span style="color:{c}">&#9632; {r}</span>'
    timeline_html = f"""
  <section class="timeline-section">
    <h2 class="section-title">&#128197; Macro Regime Timeline (24 months)</h2>
    <div class="tl-bar">{tl_segments}
      <div class="tl-marker" title="You are here">
        <div class="tl-dot"></div>
        <span class="tl-marker-label">NOW</span>
      </div>
    </div>
    <div class="tl-month-labels">{tl_month_labels}</div>
    <div class="tl-legend">{tl_legend}</div>
    <p class="tl-description">{regime_desc}</p>
  </section>"""

    # Macro map HTML container
    macro_map_html = """
      <section class="macro-map-section" id="macro-map-section" style="display:none">
        <h2 class="section-title">&#9889; Macro Map — Indicator Network</h2>
        <div id="macro-map-container"></div>
      </section>"""

    # Build summary panels row
    forward_look = forward_look or []
    fl_items = ""
    for fl in forward_look:
        fl_impacts_html = ""
        for fname, farrow, fcolor in fl.get("impacts", []):
            fl_impacts_html += (
                f'<span class="fl-impact" style="color:{fcolor}">'
                f'{fname} {farrow}</span>')
        fl_items += (
            f'<div class="fl-scenario">'
            f'<div class="fl-sc-head">'
            f'<span class="fl-sc-title">{fl["title"]}</span>'
            f'<span class="fl-sc-prob" style="color:{fl["prob_color"]}">'
            f'{fl["probability"]}</span></div>'
            f'<p class="fl-sc-text">{fl["text"]}</p>'
            f'<div class="fl-sc-impacts">{fl_impacts_html}</div>'
            f'</div>')
    if not fl_items:
        fl_items = '<p style="color:var(--muted);font-size:0.78rem;">No scenarios triggered by current conditions.</p>'

    # Extract first sentence for collapsed previews
    import re as _re
    def _first_sentence(text):
        text = str(text).strip()
        m = _re.match(r'([^.!?]+[.!?])', text)
        return m.group(1).strip() if m else text[:80]

    weekly_preview = _first_sentence(weekly_wrap)
    fl_preview = _first_sentence(fl_items.replace('<', ' <')) if fl_items else ""
    # Clean HTML tags from preview
    fl_preview = _re.sub(r'<[^>]+>', '', fl_preview).strip()
    if not fl_preview and forward_look:
        fl_preview = forward_look[0].get("title", "Scenarios available")

    summary_panels_html = f"""
  <div class="summary-row">
    <div class="summary-panel sp-daily">
      <div class="sp-header" onclick="toggleSP(this)">
        <span>&#128203; Daily Brief</span>
        <span class="sp-toggle">&#9660;</span>
      </div>
      <div class="sp-body">
        <p class="narrative-text">{narrative}</p>
        <p class="daily-brief-text">{daily_brief}</p>
        {changes_inline_html}
      </div>
    </div>
    <div class="summary-panel sp-weekly">
      <div class="sp-header" onclick="toggleSP(this)">
        <span>&#128202; Weekly Wrap</span>
        <span class="sp-toggle">&#9654;</span>
      </div>
      <div class="sp-preview">{weekly_preview}...</div>
      <div class="sp-body collapsed">{weekly_wrap}</div>
    </div>
    <div class="summary-panel sp-forward">
      <div class="sp-header" onclick="toggleSP(this)">
        <span>&#128302; Forward Look</span>
        <span class="sp-toggle">&#9654;</span>
      </div>
      <div class="sp-preview">{fl_preview}...</div>
      <div class="sp-body collapsed">{fl_items}</div>
    </div>
  </div>"""

    vix_val = indicators.get("VIX", {}).get("value")
    oil_val = indicators.get("OIL_WTI", {}).get("value")
    sp_val  = indicators.get("SP500", {}).get("value")
    vix_str = f"{vix_val:.2f}" if vix_val else "N/A"
    oil_str = f"${oil_val:.2f}" if oil_val else "N/A"
    sp_str  = f"{sp_val:,.0f}" if sp_val else "N/A"

    # Regime label
    if counts["BEARISH"] > counts["BULLISH"] + 3:
        regime_label = "RISK-OFF"
        regime_color = "#ff4d4d"
    elif counts["BULLISH"] > counts["BEARISH"] + 3:
        regime_label = "RISK-ON"
        regime_color = "#4dff91"
    else:
        regime_label = "MIXED / CAUTIOUS"
        regime_color = "#ffd24d"

    # Build section cards
    sections_html = ""
    by_section = {}
    for key, ind in indicators.items():
        cat = ind.get("category", "Other")
        by_section.setdefault(cat, []).append((key, ind))

    for section in SECTION_ORDER:
        items = by_section.get(section, [])
        if not items:
            continue
        icon = SECTION_ICONS.get(section, "")
        cards = ""
        for key, ind in items:
            sig = ind.get("signal", "NEUTRAL")
            sig_class = sig.lower()
            arrow = ind.get("direction", "▬")
            arrow_class = "up" if arrow == "▲" else ("down" if arrow == "▼" else "flat")
            val = value_display(key, ind)
            prev = prev_display(key, ind)
            pdate = prev_date_display(ind)
            commentary = ind.get("commentary", "")
            manual = "🖊 Manual Input" if ind.get("manual_input") else ""
            has_history = "true" if ind.get("history") else "false"
            rbar = range_bar_html(key, ind)
            spark = sparkline_svg(key, ind)
            momentum = ind.get("momentum")
            mtag = ""
            # Only show momentum tag when direction is not flat
            if momentum and arrow_class != "flat":
                mcls = "accel" if momentum == "ACCEL" else "decel"
                mtag = f'<span class="momentum-tag {mcls}">{momentum}</span>'
            tooltip = INDICATOR_GLOSSARY.get(key, "")
            tooltip_attr = f' data-tooltip="{tooltip}"' if tooltip else ""
            heat = ind.get("heat_score", 0)
            cards += f"""
        <div class="card" data-key="{key}" data-signal="{sig_class}" data-has-history="{has_history}" data-heat="{heat}" onclick="toggleChart(this)">
          <div class="card-header">
            <span class="card-name"{tooltip_attr}>{ind.get('label', key)}</span>
            <span class="signal-badge {sig_class}">{sig}</span>
          </div>
          <div class="card-body">
            <span class="value" contenteditable="false"
                  data-original="{val}">{val}</span>
            <span class="arrow {arrow_class}">{arrow}</span>
            {mtag}
          </div>
          {spark}
          <div class="card-prev">prev: <span class="prev-val">{prev}</span>{f' <span class="prev-date">{pdate}</span>' if pdate else ''}
            {f'<span class="manual-tag">{manual}</span>' if manual else ''}
          </div>
          {rbar}
          <div class="commentary">{commentary}</div>
          <div class="chart-panel">
            <div class="chart-periods">
              <button class="period-btn" data-period="1W" onclick="event.stopPropagation();setPeriod(this)">1W</button>
              <button class="period-btn" data-period="1M" onclick="event.stopPropagation();setPeriod(this)">1M</button>
              <button class="period-btn active" data-period="3M" onclick="event.stopPropagation();setPeriod(this)">3M</button>
              <button class="period-btn" data-period="6M" onclick="event.stopPropagation();setPeriod(this)">6M</button>
              <button class="period-btn" data-period="1Y" onclick="event.stopPropagation();setPeriod(this)">1Y</button>
            </div>
            <div class="chart-wrap"><canvas class="chart-canvas"></canvas></div>
            <div class="related-section"></div>
          </div>
        </div>"""

        sections_html += f"""
      <section class="section">
        <h2 class="section-title">{icon} {section}</h2>
        <div class="cards-grid">{cards}
        </div>
      </section>"""
        # Insert scenarios right after Market Indices section
        if section == "Market Indices":
            sections_html += scenarios_html

    # Upcoming releases with countdown
    cat_colors = {
        "Inflation": "#ff4d4d", "Labour": "#4d9fff", "Activity": "#ffd24d",
        "Rates": "#a855f7", "Sentiment": "#00e676", "Growth": "#ff6b35",
    }
    upcoming_items = ""
    for name, date_str, days, cat in upcoming_releases:
        if days == 0:
            badge = '<span class="countdown-badge today">TODAY</span>'
        elif days == 1:
            badge = '<span class="countdown-badge tomorrow">1d</span>'
        elif days <= 7:
            badge = f'<span class="countdown-badge soon">{days}d</span>'
        else:
            badge = f'<span class="countdown-badge">{days}d</span>'
        cat_color = cat_colors.get(cat, "var(--muted)")
        cat_badge = (f'<span class="cat-badge" style="color:{cat_color};'
                     f'border-color:{cat_color}44;background:{cat_color}18">'
                     f'{cat}</span>')
        upcoming_items += (
            f'<div class="upcoming-item">'
            f'<div class="rel-top">{badge}<span class="rel-date">{date_str}</span></div>'
            f'<span class="rel-name">{name}</span>'
            f'<div class="rel-bottom">{cat_badge}</div>'
            f'</div>')
    upcoming = f"""
      <section class="section upcoming">
        <h2 class="section-title">📅 Upcoming Releases</h2>
        <div class="upcoming-grid">{upcoming_items}</div>
      </section>"""

    # Narrative is now merged into Daily Brief panel (no standalone panel)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0, viewport-fit=cover">
  <title>Macro Pulse Dashboard</title>
  <!-- PWA -->
  <link rel="manifest" href="manifest.json">
  <meta name="theme-color" content="#0b0f19">
  <!-- iOS home screen -->
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Macro Pulse">
  <link rel="apple-touch-icon" href="icon-192.png">
  <link rel="apple-touch-icon" sizes="180x180" href="icon-192.png">
  <link rel="apple-touch-icon" sizes="512x512" href="icon-512.png">
  <!-- iOS splash screen fallback -->
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <style>
    /* PWA standalone: hide Safari UI chrome, add safe-area padding */
    @media all and (display-mode: standalone) {{
      body {{ padding-top: env(safe-area-inset-top); }}
      .topbar {{ padding-top: calc(env(safe-area-inset-top) + 8px); }}
    }}
  </style>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600;700&family=DM+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
  <style>
    :root {{
      --bg:       #0b0f19;
      --surface:  #131929;
      --surface2: #1c2540;
      --border:   #252e4a;
      --text:     #e8eaf2;
      --muted:    #8892b0;
      --bullish:  #00e676;
      --bearish:  #ff4d4d;
      --neutral:  #ffd24d;
      --accent:   #4d9fff;
      --font-data: 'JetBrains Mono', monospace;
      --font-ui:   'DM Sans', sans-serif;
    }}
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    html {{
      -webkit-text-size-adjust: 100%;
      scroll-behavior: smooth;
    }}
    body {{
      background: var(--bg);
      color: var(--text);
      font-family: var(--font-ui);
      min-height: 100vh;
      min-height: 100dvh;
      -webkit-tap-highlight-color: transparent;
      -webkit-touch-callout: none;
      overscroll-behavior-y: none;
      padding-bottom: env(safe-area-inset-bottom);
    }}

    /* ── TOP BAR ────────────────────────────────── */
    .topbar {{
      background: var(--surface);
      border-bottom: 1px solid var(--border);
      padding: 12px 28px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      flex-wrap: wrap;
      gap: 12px;
      position: sticky;
      top: 0;
      z-index: 100;
    }}
    .topbar-left {{ display: flex; align-items: center; gap: 20px; }}
    .logo {{ font-family: var(--font-data); font-size: 1.1rem; font-weight: 700;
             color: var(--accent); letter-spacing: 0.05em; }}
    .regime-badge {{
      padding: 4px 14px;
      border-radius: 20px;
      font-size: 0.75rem;
      font-weight: 600;
      letter-spacing: 0.08em;
      background: {regime_color}22;
      color: {regime_color};
      border: 1px solid {regime_color}55;
    }}
    .topbar-stats {{ display: flex; gap: 24px; align-items: center; }}
    .stat {{ display: flex; flex-direction: column; align-items: center; }}
    .stat-label {{ font-size: 0.65rem; color: var(--muted); text-transform: uppercase;
                   letter-spacing: 0.06em; }}
    .stat-val {{ font-family: var(--font-data); font-size: 0.95rem; font-weight: 600; }}
    .topbar-sep {{
      width: 1px;
      height: 24px;
      background: var(--border);
      margin: 0 4px;
    }}
    .view-toggles {{
      display: flex;
      align-items: center;
      gap: 8px;
    }}
    .view-label {{
      font-size: 0.68rem;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.06em;
      font-weight: 600;
    }}
    .topbar-right {{ display: flex; align-items: center; gap: 16px; }}
    .timestamp {{ font-size: 0.72rem; color: var(--muted); font-family: var(--font-data); }}
    .edit-btn {{
      padding: 6px 18px;
      border-radius: 6px;
      border: 1px solid var(--accent);
      background: transparent;
      color: var(--accent);
      font-family: var(--font-ui);
      font-size: 0.8rem;
      cursor: pointer;
      transition: all 0.2s;
    }}
    .edit-btn:hover {{ background: var(--accent); color: var(--bg); }}
    .edit-btn.active {{ background: var(--accent); color: var(--bg); }}
    .next-update {{
      font-size: 0.7rem;
      color: var(--muted);
      font-family: var(--font-ui);
      display: flex;
      flex-direction: column;
      align-items: center;
      gap: 1px;
    }}
    .next-update .next-label {{
      text-transform: uppercase;
      letter-spacing: 0.05em;
      font-size: 0.6rem;
    }}
    .next-update .next-time {{
      color: var(--accent);
      font-family: var(--font-data);
      font-size: 0.75rem;
    }}
    .reload-spinner {{
      display: inline-block;
      width: 14px; height: 14px;
      border: 2px solid var(--accent);
      border-top-color: transparent;
      border-radius: 50%;
      animation: spin 0.8s linear infinite;
      vertical-align: middle;
      margin-right: 4px;
    }}
    .toast-msg {{
      position: fixed;
      bottom: 30px;
      left: 50%;
      transform: translateX(-50%) translateY(20px);
      background: #1e1e1e;
      color: #ffd24d;
      border: 1px solid #ffd24d44;
      padding: 12px 24px;
      border-radius: 8px;
      font-family: var(--font-ui);
      font-size: 0.85rem;
      z-index: 9999;
      opacity: 0;
      transition: opacity 0.3s, transform 0.3s;
      white-space: nowrap;
    }}
    .toast-msg.show {{
      opacity: 1;
      transform: translateX(-50%) translateY(0);
    }}

    /* ── SIGNAL COUNTS ──────────────────────────── */
    .signal-summary {{
      display: flex;
      gap: 16px;
      align-items: center;
      font-family: var(--font-data);
      font-size: 0.85rem;
    }}
    .sig-count {{ padding: 3px 12px; border-radius: 12px; font-weight: 600; font-size: 0.78rem; }}
    .sig-count.bull {{ background: #00e67622; color: var(--bullish); border: 1px solid #00e67644; }}
    .sig-count.bear {{ background: #ff4d4d22; color: var(--bearish); border: 1px solid #ff4d4d44; }}
    .sig-count.neut {{ background: #ffd24d22; color: var(--neutral); border: 1px solid #ffd24d44; }}

    /* ── MAIN LAYOUT ────────────────────────────── */
    .main {{ max-width: 1400px; margin: 0 auto; padding: 28px 24px; }}
    .section {{ margin-bottom: 36px; }}
    .section-title {{
      font-size: 0.85rem;
      font-weight: 600;
      text-transform: uppercase;
      letter-spacing: 0.1em;
      color: var(--muted);
      margin-bottom: 16px;
      padding-bottom: 8px;
      border-bottom: 1px solid var(--border);
    }}

    /* ── CARDS ──────────────────────────────────── */
    .cards-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
      gap: 14px;
    }}
    .card {{
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 16px 18px;
      transition: border-color 0.2s, box-shadow 0.2s;
    }}
    .card:hover {{
      border-color: var(--accent);
      box-shadow: 0 0 0 1px var(--accent)33;
    }}
    .card-header {{
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      margin-bottom: 10px;
    }}
    .card-name {{
      font-size: 0.78rem;
      font-weight: 500;
      color: var(--muted);
      line-height: 1.3;
      max-width: 65%;
    }}
    .signal-badge {{
      font-size: 0.65rem;
      font-weight: 700;
      letter-spacing: 0.06em;
      padding: 2px 8px;
      border-radius: 4px;
    }}
    .signal-badge.bullish {{ background: #00e67622; color: var(--bullish); border: 1px solid #00e67644; }}
    .signal-badge.bearish {{ background: #ff4d4d22; color: var(--bearish); border: 1px solid #ff4d4d44; }}
    .signal-badge.neutral {{ background: #ffd24d22; color: var(--neutral); border: 1px solid #ffd24d44; }}

    .card-body {{
      display: flex;
      align-items: baseline;
      gap: 8px;
      margin-bottom: 6px;
    }}
    .value {{
      font-family: var(--font-data);
      font-size: 1.5rem;
      font-weight: 700;
      color: var(--text);
      line-height: 1;
    }}
    .value[contenteditable="true"] {{
      outline: none;
      border-bottom: 1px dashed var(--accent);
      color: var(--accent);
    }}
    .arrow {{ font-size: 1rem; font-family: var(--font-data); }}
    .arrow.up   {{ color: var(--bearish); }}   /* up = potentially bad (rates, VIX) — neutral color used */
    .arrow.down {{ color: var(--bullish); }}
    .arrow.flat {{ color: var(--muted); }}

    .card-prev {{
      font-size: 0.72rem;
      color: var(--muted);
      font-family: var(--font-data);
      margin-bottom: 8px;
    }}
    .manual-tag {{
      font-size: 0.65rem;
      color: var(--neutral);
      margin-left: 8px;
    }}
    .commentary {{
      font-size: 0.75rem;
      color: #aab0c8;
      line-height: 1.5;
      border-top: 1px solid var(--border);
      padding-top: 8px;
    }}
    .prev-date {{
      color: #667194;
      font-size: 0.68rem;
    }}

    /* ── NARRATIVE (inside Daily Brief) ──────────── */
    .narrative-text {{
      font-size: 0.82rem;
      line-height: 1.65;
      color: #c0c6db;
      margin: 0 0 10px 0;
    }}
    .daily-brief-text {{
      font-size: 0.78rem;
      line-height: 1.6;
      color: #aab0c8;
      margin: 0 0 10px 0;
    }}
    /* ── CHANGES INLINE (inside Daily Brief) ───── */
    .changes-inline {{
      border-top: 1px solid var(--border);
      margin-top: 10px;
      padding-top: 8px;
    }}
    .changes-inline-header {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      cursor: pointer;
      font-size: 0.75rem;
      font-weight: 600;
      color: var(--muted);
      padding: 4px 0;
    }}
    .changes-inline-header:hover {{ color: var(--text); }}
    /* ── COLLAPSED PREVIEW ─────────────────────── */
    .sp-preview {{
      padding: 0 16px 10px;
      font-size: 0.75rem;
      color: #667194;
      line-height: 1.4;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    /* Preview is visible by default; hidden when body is expanded via JS */

    /* ── RANGE BAR ─────────────────────────────── */
    .range-bar {{ margin: 6px 0 8px; }}
    .range-label {{
      font-size: 0.6rem;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.04em;
      margin-bottom: 3px;
    }}
    .range-track {{
      height: 4px;
      background: var(--border);
      border-radius: 2px;
      position: relative;
    }}
    .range-fill {{
      position: absolute;
      top: 0; left: 0;
      height: 100%;
      background: linear-gradient(90deg, var(--accent)44, var(--accent)88);
      border-radius: 2px;
    }}
    .range-marker {{
      position: absolute;
      top: -3px;
      width: 10px; height: 10px;
      border-radius: 50%;
      background: var(--accent);
      border: 1.5px solid var(--bg);
      transform: translateX(-50%);
    }}
    .range-labels {{
      display: flex;
      justify-content: space-between;
      font-size: 0.58rem;
      color: #667194;
      font-family: var(--font-data);
      margin-top: 3px;
    }}
    .range-labels span:nth-child(2) {{
      color: #4a5278;
      text-transform: uppercase;
      letter-spacing: 0.05em;
    }}

    /* ── COUNTDOWN BADGES ──────────────────────── */
    .countdown-badge {{
      font-family: var(--font-data);
      font-size: 0.72rem;
      font-weight: 700;
      color: var(--muted);
      background: var(--surface);
      border: 1px solid var(--border);
      padding: 2px 8px;
      border-radius: 4px;
      min-width: 36px;
      text-align: center;
    }}
    .countdown-badge.today {{
      background: var(--bearish)22;
      color: var(--bearish);
      border-color: var(--bearish)44;
    }}
    .countdown-badge.tomorrow {{
      background: var(--neutral)22;
      color: var(--neutral);
      border-color: var(--neutral)44;
    }}
    .countdown-badge.soon {{
      background: var(--accent)22;
      color: var(--accent);
      border-color: var(--accent)44;
    }}

    /* ── RELATED INDICATORS ────────────────────── */
    .related-section {{
      display: none;
      margin-top: 10px;
      padding-top: 10px;
      border-top: 1px solid var(--border);
    }}
    .card.expanded .related-section {{
      display: flex;
      align-items: center;
      gap: 8px;
      flex-wrap: wrap;
    }}
    .related-label {{
      font-size: 0.7rem;
      color: var(--muted);
      font-weight: 600;
      letter-spacing: 0.04em;
    }}
    .related-chip {{
      display: inline-flex;
      align-items: center;
      gap: 5px;
      background: var(--surface2);
      border: 1px solid var(--border);
      border-radius: 6px;
      padding: 4px 10px;
      font-size: 0.68rem;
      color: var(--text);
      font-family: var(--font-data);
      cursor: pointer;
      transition: border-color 0.2s;
    }}
    .related-chip:hover {{ border-color: var(--accent); }}
    .related-chip .rel-arrow {{ font-size: 0.7rem; }}
    .related-chip .rel-arrow.up {{ color: var(--bearish); }}
    .related-chip .rel-arrow.down {{ color: var(--bullish); }}
    .related-chip .rel-arrow.flat {{ color: var(--muted); }}
    .related-chip .rel-val {{
      color: var(--muted);
      font-size: 0.62rem;
    }}

    /* ── CHART PANEL ───────────────────────────── */
    .card {{ cursor: pointer; }}
    .chart-panel {{
      display: none;
      margin-top: 12px;
      padding-top: 12px;
      border-top: 1px solid var(--border);
    }}
    .card.expanded .chart-panel {{
      display: block;
    }}
    .card.expanded {{
      grid-column: 1 / -1;
      border-color: var(--accent);
    }}
    .chart-periods {{
      display: flex;
      gap: 6px;
      margin-bottom: 8px;
    }}
    .period-btn {{
      padding: 3px 10px;
      border-radius: 4px;
      border: 1px solid var(--border);
      background: transparent;
      color: var(--muted);
      font-family: var(--font-data);
      font-size: 0.7rem;
      cursor: pointer;
      transition: all 0.2s;
    }}
    .period-btn:hover, .period-btn.active {{
      background: var(--accent);
      color: var(--bg);
      border-color: var(--accent);
    }}
    .chart-wrap {{
      position: relative;
      height: 220px;
    }}

    /* ── UPCOMING ──────────────────────────────── */
    .upcoming-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(180px, 1fr));
      gap: 12px;
    }}
    .upcoming-item {{
      background: var(--surface2);
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 12px 14px;
      display: flex;
      flex-direction: column;
      gap: 8px;
    }}
    .rel-top {{
      display: flex;
      align-items: center;
      justify-content: space-between;
    }}
    .rel-date {{ font-size: 0.68rem; color: var(--muted); font-family: var(--font-data); }}
    .rel-name {{ font-size: 0.8rem; font-weight: 500; line-height: 1.3; }}
    .rel-bottom {{ display: flex; }}
    .cat-badge {{
      font-size: 0.6rem;
      font-weight: 600;
      letter-spacing: 0.04em;
      padding: 1px 7px;
      border-radius: 3px;
      border: 1px solid;
      white-space: nowrap;
    }}

    /* ── EDIT MODE ──────────────────────────────── */
    body.edit-mode .value {{ cursor: text; }}
    body.edit-mode .card {{ border-color: var(--accent)44; }}

    /* ── SPARKLINE ──────────────────────────────── */
    .sparkline {{ display: block; margin: 8px 0; }}

    /* ── MOMENTUM TAG ──────────────────────────── */
    .momentum-tag {{
      font-size: 0.6rem;
      font-weight: 700;
      letter-spacing: 0.06em;
      padding: 1px 6px;
      border-radius: 3px;
      font-family: var(--font-data);
      vertical-align: middle;
    }}
    .momentum-tag.accel {{
      background: #4d9fff33;
      color: #6ab0ff;
    }}
    .momentum-tag.decel {{
      background: #8892b033;
      color: #667194;
    }}

    /* ── HEAT MAP ──────────────────────────────── */
    .heatmap-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(130px, 1fr));
      gap: 6px;
    }}
    .hm-cell {{
      padding: 10px 12px;
      border-radius: 8px;
      cursor: pointer;
      transition: transform 0.15s, box-shadow 0.15s;
      text-align: center;
    }}
    .hm-cell:hover {{
      transform: scale(1.04);
      box-shadow: 0 0 12px rgba(255,255,255,0.1);
    }}
    .hm-label {{
      display: block;
      font-size: 0.68rem;
      color: rgba(255,255,255,0.85);
      font-weight: 500;
      margin-bottom: 4px;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .hm-val {{
      display: block;
      font-family: var(--font-data);
      font-size: 0.82rem;
      font-weight: 700;
      color: #fff;
    }}
    .hm-legend {{
      display: flex;
      gap: 16px;
      justify-content: center;
      margin-top: 12px;
      font-size: 0.7rem;
      font-family: var(--font-data);
    }}
    .heatmap-btn {{
      padding: 6px 18px;
      border-radius: 6px;
      border: 1px solid var(--accent);
      background: transparent;
      color: var(--accent);
      font-family: var(--font-ui);
      font-size: 0.8rem;
      cursor: pointer;
      transition: all 0.2s;
    }}
    .heatmap-btn:hover {{ background: var(--accent); color: var(--bg); }}
    .heatmap-btn.active {{ background: var(--accent); color: var(--bg); }}

    /* ── CHANGES (inline in Daily Brief) ──────── */
    .changes-toggle {{ color: var(--muted); font-size: 0.7rem; transition: transform 0.2s; }}
    .changes-body {{
      padding: 0 18px 14px;
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
      gap: 8px;
    }}
    .changes-body.collapsed {{ display: none; }}
    .change-item {{
      display: flex;
      align-items: center;
      gap: 10px;
      padding: 8px 12px;
      background: var(--surface2);
      border-radius: 6px;
      border: 1px solid var(--border);
    }}
    .change-name {{
      font-size: 0.78rem;
      font-weight: 500;
      color: var(--text);
      min-width: 100px;
    }}
    .change-vals {{
      font-family: var(--font-data);
      font-size: 0.75rem;
      color: var(--accent);
    }}
    .change-ctx {{
      font-size: 0.7rem;
      color: var(--muted);
      margin-left: auto;
      white-space: nowrap;
    }}

    /* ── GLOSSARY TOOLTIP ──────────────────────── */
    [data-tooltip] {{
      position: relative;
      cursor: help;
    }}
    [data-tooltip]:hover::after {{
      content: attr(data-tooltip);
      position: absolute;
      bottom: calc(100% + 8px);
      left: 50%;
      transform: translateX(-50%);
      background: #1c2540;
      color: #c0c6db;
      border: 1px solid var(--border);
      border-radius: 6px;
      padding: 8px 12px;
      font-size: 0.72rem;
      line-height: 1.4;
      width: max-content;
      max-width: 260px;
      z-index: 200;
      pointer-events: none;
      white-space: normal;
      font-weight: 400;
      font-family: var(--font-ui);
      box-shadow: 0 4px 16px rgba(0,0,0,0.3);
    }}
    [data-tooltip]:hover::before {{
      content: '';
      position: absolute;
      bottom: calc(100% + 2px);
      left: 50%;
      transform: translateX(-50%);
      border: 5px solid transparent;
      border-top-color: #1c2540;
      z-index: 201;
      pointer-events: none;
    }}
    .glossary-term {{
      border-bottom: 1px dotted #667194;
      cursor: help;
      position: relative;
    }}

    /* ── SUMMARY PANELS ROW ──────────────────── */
    .summary-row {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 14px;
      margin-bottom: 24px;
      border-bottom: 1px solid var(--border);
      padding-bottom: 24px;
    }}
    .summary-panel {{
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 10px;
      overflow: hidden;
    }}
    .sp-daily  {{ border-left: 3px solid var(--accent); }}
    .sp-weekly {{ border-left: 3px solid #ffd24d; }}
    .sp-forward {{ border-left: 3px solid #a855f7; }}
    .sp-header {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      padding: 10px 16px;
      cursor: pointer;
      font-size: 0.82rem;
      font-weight: 600;
      color: var(--text);
      text-transform: uppercase;
      letter-spacing: 0.06em;
    }}
    .sp-header:hover {{ background: var(--surface2); }}
    .sp-toggle {{
      color: var(--muted);
      font-size: 0.7rem;
      transition: transform 0.2s;
    }}
    .sp-body {{
      padding: 0 16px 14px;
      font-size: 0.78rem;
      color: #aab0c8;
      line-height: 1.6;
    }}
    .sp-body.collapsed {{
      display: none;
    }}
    /* Forward Look scenario sub-cards */
    .fl-scenario {{
      background: var(--surface2);
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 10px 14px;
      margin-bottom: 8px;
    }}
    .fl-scenario:last-child {{ margin-bottom: 0; }}
    .fl-sc-head {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-bottom: 6px;
    }}
    .fl-sc-title {{
      font-size: 0.78rem;
      font-weight: 600;
      color: var(--text);
    }}
    .fl-sc-prob {{
      font-family: var(--font-data);
      font-size: 0.65rem;
      font-weight: 700;
      padding: 1px 7px;
      border-radius: 3px;
      background: var(--surface);
      letter-spacing: 0.04em;
    }}
    .fl-sc-text {{
      font-size: 0.75rem;
      color: #aab0c8;
      line-height: 1.55;
      margin: 0 0 6px 0;
    }}
    .fl-sc-impacts {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
    }}
    .fl-impact {{
      font-family: var(--font-data);
      font-size: 0.65rem;
      font-weight: 600;
    }}
    @media (max-width: 900px) {{
      .summary-row {{
        grid-template-columns: 1fr;
      }}
      .sp-body {{ display: none; }}
    }}

    /* ── MACRO MAP ─────────────────────────────── */
    .macro-map-section {{
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 16px;
      margin-bottom: 28px;
    }}
    #macro-map-container {{
      width: 100%;
      min-height: 600px;
      overflow: hidden;
    }}
    #macro-map-container svg {{
      display: block;
      margin: 0 auto;
    }}
    .macromap-btn {{
      padding: 6px 18px;
      border-radius: 6px;
      border: 1px solid var(--accent);
      background: transparent;
      color: var(--accent);
      font-family: var(--font-ui);
      font-size: 0.8rem;
      cursor: pointer;
      transition: all 0.2s;
    }}
    .macromap-btn:hover {{ background: var(--accent); color: var(--bg); }}
    .macromap-btn.active {{ background: var(--accent); color: var(--bg); }}

    /* ── SCENARIOS ──────────────────────────────── */
    .scenarios-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(320px, 1fr));
      gap: 14px;
    }}
    .scenario-box {{
      background: var(--surface);
      border: 1px solid var(--border);
      border-left: 3px solid;
      border-radius: 10px;
      padding: 16px 18px;
    }}
    .sc-header {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-bottom: 10px;
    }}
    .sc-title {{
      font-size: 0.9rem;
      font-weight: 600;
      color: var(--text);
    }}
    .sc-prob {{
      font-family: var(--font-data);
      font-size: 0.72rem;
      font-weight: 700;
      padding: 2px 8px;
      border-radius: 4px;
      background: var(--surface2);
      letter-spacing: 0.04em;
    }}
    .sc-content {{
      font-size: 0.78rem;
      color: #aab0c8;
      line-height: 1.6;
      margin-bottom: 10px;
    }}
    .sc-impacts {{
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      border-top: 1px solid var(--border);
      padding-top: 8px;
    }}
    .sc-impact {{
      font-family: var(--font-data);
      font-size: 0.7rem;
      font-weight: 600;
    }}

    /* ── REGIME TIMELINE ───────────────────────── */
    .timeline-section {{
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 16px 20px;
      margin-top: 28px;
    }}
    .tl-bar {{
      display: flex;
      align-items: stretch;
      height: 44px;
      border-radius: 6px;
      overflow: visible;
      position: relative;
      margin-bottom: 24px;
    }}
    .tl-segment {{
      flex: 1;
      display: flex;
      align-items: center;
      justify-content: center;
      min-width: 0;
      transition: filter 0.15s;
      position: relative;
    }}
    .tl-segment:hover {{
      filter: brightness(1.3);
      z-index: 2;
    }}
    .tl-month {{
      font-family: var(--font-data);
      font-size: 0.55rem;
      color: rgba(255,255,255,0.7);
      overflow: hidden;
      text-overflow: clip;
      white-space: nowrap;
    }}
    .tl-month-labels {{
      display: flex;
      margin-top: 4px;
    }}
    .tl-month-labels span {{
      flex: 1;
      text-align: center;
      font-family: var(--font-data);
      font-size: 0.55rem;
      color: var(--muted);
    }}
    .tl-marker {{
      position: absolute;
      right: 0;
      bottom: -20px;
      transform: translateX(50%);
      z-index: 5;
      display: flex;
      flex-direction: column;
      align-items: center;
      gap: 3px;
    }}
    .tl-dot {{
      width: 10px;
      height: 10px;
      border-radius: 50%;
      background: var(--accent);
      animation: tlPulse 1.5s ease-in-out infinite;
    }}
    .tl-marker-label {{
      font-family: var(--font-data);
      font-size: 0.55rem;
      color: var(--accent);
      white-space: nowrap;
      letter-spacing: 0.04em;
    }}
    @keyframes tlPulse {{
      0%, 100% {{ box-shadow: 0 0 0 0 rgba(77, 159, 255, 0.6); }}
      50% {{ box-shadow: 0 0 0 6px rgba(77, 159, 255, 0); }}
    }}
    .tl-legend {{
      display: flex;
      gap: 14px;
      flex-wrap: wrap;
      font-size: 0.7rem;
      font-family: var(--font-data);
      margin-top: 10px;
      margin-bottom: 8px;
    }}
    .tl-description {{
      font-size: 0.82rem;
      color: #c0c6db;
      line-height: 1.6;
      margin-top: 8px;
    }}

    /* ── FOOTER ────────────────────────────────── */
    .site-footer {{
      border-top: 1px solid var(--border);
      padding: 20px 28px;
      text-align: center;
    }}
    .footer-inner {{
      max-width: 1400px;
      margin: 0 auto;
      display: flex;
      flex-direction: column;
      gap: 4px;
    }}
    .footer-brand {{
      font-size: 0.72rem;
      color: var(--muted);
      font-weight: 500;
    }}
    .footer-meta {{
      font-size: 0.65rem;
      color: #667194;
      font-family: var(--font-data);
    }}
    .footer-cmd {{
      font-size: 0.62rem;
      color: #4a5278;
      margin-top: 4px;
    }}
    .footer-cmd code {{
      font-family: var(--font-data);
      background: var(--surface2);
      padding: 2px 8px;
      border-radius: 4px;
      border: 1px solid var(--border);
      font-size: 0.62rem;
    }}

    /* ── RESPONSIVE ─────────────────────────────── */
    @media (max-width: 600px) {{
      .topbar {{
        padding: 10px 12px;
        flex-direction: column;
        align-items: flex-start;
        gap: 8px;
      }}
      .topbar-left {{
        flex-wrap: wrap;
        gap: 8px;
        width: 100%;
      }}
      .topbar-stats {{
        gap: 8px;
        flex-wrap: wrap;
      }}
      .topbar-sep {{ display: none; }}
      .topbar-right {{
        width: 100%;
        justify-content: space-between;
        gap: 8px;
      }}
      .view-toggles {{ flex-wrap: wrap; gap: 4px; }}
      .signal-summary {{ gap: 8px; flex-wrap: wrap; }}
      .sig-count {{ font-size: 0.68rem; padding: 2px 8px; }}

      .main {{ padding: 12px 10px; }}
      .cards-grid {{
        grid-template-columns: 1fr;
        gap: 10px;
      }}
      .value {{ font-size: 1.25rem; }}
      .card {{ padding: 14px; }}
      .card-name {{ font-size: 0.75rem; max-width: 60%; }}

      .section-title {{ font-size: 0.78rem; }}
      .commentary {{ font-size: 0.73rem; }}
      .narrative-text {{ font-size: 0.78rem; }}
      .daily-brief-text {{ font-size: 0.75rem; }}

      .edit-btn {{ padding: 5px 12px; font-size: 0.75rem; }}
      .logo {{ font-size: 0.95rem; }}
      .timestamp {{ font-size: 0.65rem; }}
      .stat-val {{ font-size: 0.82rem; }}
      .stat-label {{ font-size: 0.6rem; }}

      .summary-row {{ grid-template-columns: 1fr; }}

      .macro-map-section {{ padding: 10px; overflow-x: auto; }}

      .site-footer {{ padding: 16px 10px; }}
      .footer-inner {{ flex-direction: column; gap: 6px; text-align: center; }}

      /* Safe area insets for notched devices */
      body {{
        padding-top: env(safe-area-inset-top);
        padding-left: env(safe-area-inset-left);
        padding-right: env(safe-area-inset-right);
      }}
    }}

    @media (max-width: 375px) {{
      .topbar {{ padding: 8px 8px; }}
      .main {{ padding: 10px 6px; }}
      .logo {{ font-size: 0.85rem; }}
      .value {{ font-size: 1.1rem; }}
      .card {{ padding: 12px 10px; }}
      .stat {{ min-width: 0; }}
      .regime-badge {{ font-size: 0.68rem; padding: 3px 10px; }}
    }}
  </style>
</head>
<body>

<header class="topbar">
  <div class="topbar-left">
    <span class="logo">&#9889; MACRO PULSE</span>
    <span class="regime-badge">{regime_label}</span>
    <div class="signal-summary">
      <span class="sig-count bull">&#9650; {counts['BULLISH']} BULLISH</span>
      <span class="sig-count bear">&#9660; {counts['BEARISH']} BEARISH</span>
      <span class="sig-count neut">&#9644; {counts['NEUTRAL']} NEUTRAL</span>
    </div>
    <span class="topbar-sep"></span>
    <div class="view-toggles">
      <span class="view-label">View:</span>
      <button class="heatmap-btn" onclick="toggleHeatmap(this)">&#9638; Heat Map</button>
      <button class="macromap-btn" onclick="toggleMacroMap(this)">&#9889; Macro Map</button>
    </div>
  </div>
  <div class="topbar-stats">
    <div class="stat">
      <span class="stat-label">VIX</span>
      <span class="stat-val" style="color: {'var(--bearish)' if vix_val and vix_val > 20 else 'var(--bullish)'}">{vix_str}</span>
    </div>
    <div class="stat">
      <span class="stat-label">WTI Oil</span>
      <span class="stat-val">{oil_str}</span>
    </div>
    <div class="stat">
      <span class="stat-label">S&amp;P 500</span>
      <span class="stat-val">{sp_str}</span>
    </div>
  </div>
  <div class="topbar-right">
    <span class="timestamp">Data as of: {timestamp}</span>
    <span class="topbar-sep"></span>
    <span class="next-update" id="nextUpdate"></span>
    <button class="edit-btn" onclick="toggleEditMode(this)">&#9998; Edit Mode</button>
  </div>
</header>

<main class="main">
{summary_panels_html}
{heatmap_html}
{macro_map_html}
<div id="cards-view">
{sections_html}
{upcoming}
</div>
{timeline_html}
</main>

<script src="https://cdnjs.cloudflare.com/ajax/libs/d3/7.9.0/d3.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script>
  // ── Data embedded from Python ──────────────────────────────────────────
  const HISTORY_DATA = {history_json};
  const IND_META = {ind_meta_json};
  const CORRELATIONS = {correlations_json};
  const HEAT_DATA = {heat_json};
  const MACRO_MAP_NODES = {map_nodes_json};
  const MACRO_MAP_EDGES = {map_edges_json};
  const REGIME_TIMELINE = {timeline_json};

  // Auto-refresh schedule (UTC): 13:00 (pre-market) and 21:30 (post-close)
  function updateNextRefresh() {{
    const el = document.getElementById('nextUpdate');
    if (!el) return;
    const now = new Date();
    const utcH = now.getUTCHours(), utcM = now.getUTCMinutes();
    const mins = utcH * 60 + utcM;
    // Schedule: 13:00 UTC and 21:30 UTC
    const slots = [13 * 60, 21 * 60 + 30];
    let nextMins = slots.find(s => s > mins);
    let label;
    if (nextMins !== undefined) {{
      const diff = nextMins - mins;
      const h = Math.floor(diff / 60);
      const m = diff % 60;
      label = h > 0 ? h + 'h ' + m + 'm' : m + 'm';
    }} else {{
      // Next is tomorrow morning
      const diff = (24 * 60 - mins) + slots[0];
      const h = Math.floor(diff / 60);
      const m = diff % 60;
      label = h + 'h ' + m + 'm';
    }}
    el.innerHTML = '<span class="next-label">Next update</span><span class="next-time">' + label + '</span>';
  }}
  updateNextRefresh();
  setInterval(updateNextRefresh, 60000);

  function showToast(msg) {{
    const t = document.createElement('div');
    t.className = 'toast-msg';
    t.textContent = msg;
    document.body.appendChild(t);
    requestAnimationFrame(() => t.classList.add('show'));
    setTimeout(() => {{
      t.classList.remove('show');
      setTimeout(() => t.remove(), 400);
    }}, 4000);
  }}

  // ── Chart instances registry ─────────────────────────────────────────────
  const chartInstances = {{}};

  function getSignalColor(card) {{
    const sig = card.dataset.signal;
    if (sig === 'bullish') return '#00e676';
    if (sig === 'bearish') return '#ff4d4d';
    return '#ffd24d';
  }}

  function filterByPeriod(data, period) {{
    if (!data || data.length === 0) return data;
    const now = new Date();
    const cutoff = new Date();
    switch(period) {{
      case '1W': cutoff.setDate(now.getDate() - 7); break;
      case '1M': cutoff.setMonth(now.getMonth() - 1); break;
      case '3M': cutoff.setMonth(now.getMonth() - 3); break;
      case '6M': cutoff.setMonth(now.getMonth() - 6); break;
      case '1Y': cutoff.setFullYear(now.getFullYear() - 1); break;
    }}
    const cutoffStr = cutoff.toISOString().split('T')[0];
    return data.filter(d => d[0] >= cutoffStr);
  }}

  function formatDateLabel(dateStr) {{
    const d = new Date(dateStr + 'T00:00:00');
    const months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
    return d.getDate() + ' ' + months[d.getMonth()];
  }}

  function renderChart(card, period) {{
    const key = card.dataset.key;
    const data = HISTORY_DATA[key];
    if (!data || data.length === 0) return;

    const filtered = filterByPeriod(data, period);
    if (filtered.length === 0) return;

    const canvas = card.querySelector('.chart-canvas');
    const ctx = canvas.getContext('2d');
    const color = getSignalColor(card);

    if (chartInstances[key]) {{
      chartInstances[key].destroy();
    }}

    chartInstances[key] = new Chart(ctx, {{
      type: 'line',
      data: {{
        labels: filtered.map(d => d[0]),
        datasets: [{{
          data: filtered.map(d => d[1]),
          borderColor: color,
          backgroundColor: color + '1a',
          borderWidth: 1.5,
          pointRadius: 0,
          pointHoverRadius: 4,
          pointHoverBackgroundColor: color,
          fill: true,
          tension: 0.3,
        }}]
      }},
      options: {{
        responsive: true,
        maintainAspectRatio: false,
        plugins: {{
          legend: {{ display: false }},
          tooltip: {{
            mode: 'index',
            intersect: false,
            backgroundColor: '#1c2540',
            titleColor: '#e8eaf2',
            bodyColor: '#e8eaf2',
            borderColor: '#252e4a',
            borderWidth: 1,
            titleFont: {{ family: "'JetBrains Mono', monospace", size: 11 }},
            bodyFont: {{ family: "'JetBrains Mono', monospace", size: 11 }},
            callbacks: {{
              title: function(items) {{
                return formatDateLabel(items[0].label);
              }},
              label: function(ctx) {{
                return '  ' + ctx.parsed.y.toLocaleString(undefined, {{
                  minimumFractionDigits: 2, maximumFractionDigits: 2
                }});
              }}
            }}
          }}
        }},
        scales: {{
          x: {{
            grid: {{ color: '#1f2937', drawBorder: false }},
            ticks: {{
              color: '#8892b0',
              maxTicksLimit: 6,
              font: {{ size: 10, family: "'JetBrains Mono', monospace" }},
              callback: function(val, idx) {{
                return formatDateLabel(this.getLabelForValue(val));
              }}
            }},
          }},
          y: {{
            grid: {{ color: '#1f2937', drawBorder: false }},
            ticks: {{
              color: '#8892b0',
              font: {{ size: 10, family: "'JetBrains Mono', monospace" }},
            }},
          }}
        }},
        interaction: {{
          mode: 'nearest',
          axis: 'x',
          intersect: false,
        }}
      }}
    }});
  }}

  function toggleChart(card) {{
    // Don't toggle charts when in edit mode
    if (document.body.classList.contains('edit-mode')) return;
    // Don't toggle if no history data
    if (card.dataset.hasHistory === 'false') return;

    const wasExpanded = card.classList.contains('expanded');

    // Collapse all expanded cards
    document.querySelectorAll('.card.expanded').forEach(c => {{
      c.classList.remove('expanded');
      const k = c.dataset.key;
      if (chartInstances[k]) {{
        chartInstances[k].destroy();
        delete chartInstances[k];
      }}
    }});

    if (!wasExpanded) {{
      card.classList.add('expanded');
      // Reset period buttons to 3M
      card.querySelectorAll('.period-btn').forEach(btn => {{
        btn.classList.toggle('active', btn.dataset.period === '3M');
      }});
      // Small delay so container is visible before chart renders
      requestAnimationFrame(() => renderChart(card, '3M'));
      // Populate related indicators
      populateRelated(card);
    }}
  }}

  function populateRelated(card) {{
    const key = card.dataset.key;
    const section = card.querySelector('.related-section');
    if (!section) return;
    const related = CORRELATIONS[key] || [];
    if (related.length === 0) {{ section.innerHTML = ''; return; }}
    let html = '<span class="related-label">Related:</span>';
    related.forEach(rk => {{
      const meta = IND_META[rk];
      if (!meta) return;
      const dir = meta.direction;
      const arrowCls = dir === '\u25b2' ? 'up' : (dir === '\u25bc' ? 'down' : 'flat');
      html += `<span class="related-chip" onclick="event.stopPropagation();scrollToCard('${{rk}}')">`
            + `<span class="rel-arrow ${{arrowCls}}">${{dir}}</span>`
            + `${{meta.label}}`
            + `<span class="rel-val">${{meta.value}}</span>`
            + `</span>`;
    }});
    section.innerHTML = html;
  }}

  function scrollToCard(key) {{
    const card = document.querySelector(`.card[data-key="${{key}}"]`);
    if (!card) return;
    // Collapse current, expand target
    document.querySelectorAll('.card.expanded').forEach(c => {{
      c.classList.remove('expanded');
      const k = c.dataset.key;
      if (chartInstances[k]) {{ chartInstances[k].destroy(); delete chartInstances[k]; }}
    }});
    card.scrollIntoView({{ behavior: 'smooth', block: 'center' }});
    setTimeout(() => {{
      card.classList.add('expanded');
      card.querySelectorAll('.period-btn').forEach(btn => {{
        btn.classList.toggle('active', btn.dataset.period === '3M');
      }});
      requestAnimationFrame(() => renderChart(card, '3M'));
      populateRelated(card);
    }}, 400);
  }}

  function setPeriod(btn) {{
    const card = btn.closest('.card');
    const period = btn.dataset.period;
    card.querySelectorAll('.period-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    renderChart(card, period);
  }}

  // ── Edit mode ────────────────────────────────────────────────────────────
  function toggleEditMode(btn) {{
    document.body.classList.toggle('edit-mode');
    const isEdit = document.body.classList.contains('edit-mode');
    btn.textContent = isEdit ? '✓ Save Changes' : '✎ Edit Mode';
    btn.classList.toggle('active', isEdit);
    document.querySelectorAll('.value').forEach(el => {{
      el.contentEditable = isEdit ? 'true' : 'false';
    }});
    if (!isEdit) {{
      const overrides = {{}};
      document.querySelectorAll('.card').forEach(card => {{
        const key = card.dataset.key;
        const valEl = card.querySelector('.value');
        if (valEl && valEl.textContent !== valEl.dataset.original) {{
          overrides[key] = valEl.textContent;
        }}
      }});
      if (Object.keys(overrides).length > 0) {{
        console.log('Manual overrides:', overrides);
        const toast = document.createElement('div');
        toast.style.cssText = 'position:fixed;bottom:24px;right:24px;background:#1c2540;border:1px solid #4d9fff;color:#e8eaf2;padding:12px 20px;border-radius:8px;font-size:0.8rem;z-index:999;';
        toast.textContent = `${{Object.keys(overrides).length}} value(s) overridden in-browser.`;
        document.body.appendChild(toast);
        setTimeout(() => toast.remove(), 4000);
      }}
    }}
  }}

  document.addEventListener('keydown', e => {{
    if ((e.ctrlKey || e.metaKey) && e.key === 'e') {{
      e.preventDefault();
      document.querySelector('.edit-btn').click();
    }}
  }});

  // ── Heat Map Toggle ─────────────────────────────────────────────────────
  function toggleHeatmap(btn) {{
    const hm = document.getElementById('heatmap-section');
    const cv = document.getElementById('cards-view');
    const mm = document.getElementById('macro-map-section');
    const mmBtn = document.querySelector('.macromap-btn');
    const isActive = btn.classList.contains('active');
    // Deactivate macro map if active
    if (!isActive && mmBtn && mmBtn.classList.contains('active')) {{
      mmBtn.classList.remove('active');
      mmBtn.textContent = '\u26a1 Macro Map';
      if (mm) mm.style.display = 'none';
    }}
    btn.classList.toggle('active', !isActive);
    btn.textContent = isActive ? '\u25a6 Heat Map' : '\u2715 Cards View';
    hm.style.display = isActive ? 'none' : 'block';
    cv.style.display = isActive ? 'block' : 'none';
    if (mm) mm.style.display = 'none';
  }}

  function switchToCard(key) {{
    const btn = document.querySelector('.heatmap-btn');
    if (btn && btn.classList.contains('active')) {{
      toggleHeatmap(btn);
    }}
    setTimeout(() => scrollToCard(key), 200);
  }}

  // ── Changes Panel Toggle ────────────────────────────────────────────────
  function toggleChanges() {{
    const body = document.getElementById('changes-body');
    const toggle = document.getElementById('changes-toggle');
    if (!body) return;
    body.classList.toggle('collapsed');
    toggle.textContent = body.classList.contains('collapsed') ? '\u25b6' : '\u25bc';
  }}

  // ── Summary Panel Toggle ────────────────────────────────────────────────
  function toggleSP(header) {{
    const panel = header.closest('.summary-panel');
    const body = panel.querySelector('.sp-body');
    const preview = panel.querySelector('.sp-preview');
    const toggle = header.querySelector('.sp-toggle');
    if (!body) return;
    body.classList.toggle('collapsed');
    const isCollapsed = body.classList.contains('collapsed');
    toggle.textContent = isCollapsed ? '\u25b6' : '\u25bc';
    if (preview) preview.style.display = isCollapsed ? 'block' : 'none';
  }}

  // ── Glossary Term Wrapping ──────────────────────────────────────────────
  const TERM_GLOSSARY = {{
    "200-day SMA": "Average price over 200 trading days \u2014 key long-term trend indicator",
    "50-day SMA": "Average price over 50 trading days \u2014 short-term trend indicator",
    "SMA": "Simple Moving Average \u2014 average closing price over N days",
    "YoY": "Year-over-Year \u2014 comparing to the same period last year",
    "MoM": "Month-over-Month \u2014 change from the previous month",
    "bps": "Basis points \u2014 1/100th of a percentage point (100bps = 1%)",
    "yield curve": "Graph of yields across maturities \u2014 shape signals economic expectations",
    "inverted": "Short-term yields above long-term \u2014 recession warning signal",
    "steepening": "Gap between long and short-term yields widening \u2014 typically bullish",
    "risk-off": "Investors fleeing to safe assets like bonds and gold",
    "risk-on": "Investors favoring stocks and higher-yielding assets",
    "flight-to-safety": "Rapid shift into safe-haven assets during market stress",
    "hawkish": "Favoring tighter monetary policy (higher rates) to fight inflation",
    "dovish": "Favoring looser monetary policy (lower rates) to support growth",
    "restrictive": "Rates above neutral \u2014 actively slowing economic growth",
    "accommodative": "Rates below neutral \u2014 actively supporting economic growth",
    "disinflation": "Inflation rate declining but still positive",
    "stagflation": "Simultaneous high inflation and weak growth",
    "Sahm Rule": "Recession trigger: 3-month avg unemployment rises 0.5%+ from its low",
    "term premium": "Extra yield for holding longer-term bonds vs rolling short-term",
    "contraction": "Economic shrinkage \u2014 declining output and employment",
    "expansion": "Economic growth \u2014 rising output and employment",
    "capitulation": "Panic selling \u2014 often marks market bottoms",
    "complacency": "Excessive calm that can precede sudden corrections",
    "goldilocks": "Moderate growth without excessive inflation",
    "contrarian": "Trading against prevailing sentiment",
    "FOMC": "Federal Open Market Committee \u2014 sets US interest rate policy",
    "OPEC": "Organization of Petroleum Exporting Countries \u2014 oil supply cartel",
    "GFC": "Global Financial Crisis of 2008",
  }};

  function wrapGlossaryTerms() {{
    const terms = Object.keys(TERM_GLOSSARY).sort((a, b) => b.length - a.length);
    document.querySelectorAll('.commentary').forEach(el => {{
      let html = el.innerHTML;
      terms.forEach(term => {{
        const escaped = term.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&');
        const regex = new RegExp('\\\\b(' + escaped + ')(?![^<]*>)', 'gi');
        html = html.replace(regex, (match) => {{
          return '<span class="glossary-term" data-tooltip="' + TERM_GLOSSARY[term] + '">' + match + '</span>';
        }});
      }});
      el.innerHTML = html;
    }});
  }}
  wrapGlossaryTerms();

  // ── D3 Macro Map + toggleMacroMap (injected as regular string) ──────────
  {D3_MACRO_MAP_JS}
</script>

<footer class="site-footer">
  <div class="footer-inner">
    <span class="footer-brand">Macro Pulse v3.0 &mdash; Built for David</span>
    <span class="footer-meta">Data: FRED, Yahoo Finance &nbsp;|&nbsp; Last refresh: {timestamp}</span>
    <span class="footer-cmd">&#8635; Refresh: <code>cd ~/macro-pulse &amp;&amp; python3 refresh_macro.py</code></span>
  </div>
</footer>

<script>
if ('serviceWorker' in navigator) {{
  navigator.serviceWorker.register('sw.js').catch(function() {{}});
}}
</script>

</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# 7. EXCEL GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

# ── Colour constants ──────────────────────────────────────────────────────────
XL_BG        = "0B0F19"
XL_SURFACE   = "131929"
XL_SURFACE2  = "1C2540"
XL_TEXT      = "E8EAF2"
XL_MUTED     = "8892B0"
XL_BULLISH   = "00E676"
XL_BEARISH   = "FF4D4D"
XL_NEUTRAL   = "FFD24D"
XL_ACCENT    = "4D9FFF"
XL_BORDER    = "252E4A"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color: str = XL_TEXT, bold=False, size=10, name="Calibri") -> Font:
    return Font(color=hex_color, bold=bold, size=size, name=name)


def _border_thin() -> Border:
    side = Side(border_style="thin", color=XL_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_excel(indicators: dict, timestamp: str):
    path = config.EXCEL_OUTPUT

    # Load or create
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet; we'll create ours
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── Sheet 1: Dashboard ───────────────────────────────────────────────────
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard", 0)

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = XL_ACCENT

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"⚡ MACRO PULSE  |  {timestamp}"
    title_cell.font = _font(XL_ACCENT, bold=True, size=13)
    title_cell.fill = _fill(XL_SURFACE2)
    title_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
    ws.row_dimensions[1].height = 28

    # Header row
    headers = ["Indicator", "Current", "Previous", "Change",
               "Dir", "Signal", "Commentary", "Source"]
    col_widths = [30, 14, 14, 12, 5, 10, 60, 10]
    ws.row_dimensions[2].height = 20
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _font(XL_ACCENT, bold=True, size=9)
        c.fill = _fill(XL_SURFACE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border_thin()
        _set_col_width(ws, i, w)

    row = 3
    for section in SECTION_ORDER:
        # Section header
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1,
                    value=f"  {SECTION_ICONS.get(section,'')}  {section.upper()}")
        c.font = _font(XL_MUTED, bold=True, size=8)
        c.fill = _fill(XL_SURFACE2)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18
        row += 1

        for key, ind in indicators.items():
            if ind.get("category") != section:
                continue
            sig = ind.get("signal", "NEUTRAL")
            sig_color = (XL_BULLISH if sig == "BULLISH"
                         else XL_BEARISH if sig == "BEARISH"
                         else XL_NEUTRAL)

            v = ind.get("value")
            prev = ind.get("previous")
            change = (v - prev) if (v is not None and prev is not None) else None
            vals = [
                ind.get("label", key),
                value_display(key, ind),
                prev_display(key, ind),
                f"{change:+.3f}" if change is not None else "—",
                ind.get("direction", "▬"),
                sig,
                ind.get("commentary", ""),
                ind.get("source", ""),
            ]

            ws.row_dimensions[row].height = 16
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = _fill(XL_SURFACE if row % 2 == 0 else XL_BG)
                c.alignment = Alignment(horizontal="left" if col in (1, 7) else "center",
                                        vertical="center", wrap_text=(col == 7))
                c.border = _border_thin()
                if col == 6:  # Signal column
                    c.font = Font(color=sig_color, bold=True, size=9,
                                  name="Calibri")
                elif col == 5:  # Direction
                    arrow = val
                    color = (XL_BEARISH if arrow == "▲" else
                             XL_BULLISH if arrow == "▼" else XL_MUTED)
                    c.font = Font(color=color, bold=True, size=10)
                elif col == 2:  # Current value — blue (editable)
                    c.font = _font(XL_ACCENT, size=10)
                else:
                    c.font = _font(XL_TEXT if col == 1 else XL_MUTED, size=9)
            row += 1

    # ── Sheet 2: History Log ─────────────────────────────────────────────────
    if "History Log" not in wb.sheetnames:
        ws_hist = wb.create_sheet("History Log")
        ws_hist.sheet_properties.tabColor = XL_NEUTRAL
        ws_hist.sheet_view.showGridLines = False

        # Build header from all indicator keys
        hist_headers = ["Timestamp"] + list(indicators.keys())
        for i, h in enumerate(hist_headers, 1):
            c = ws_hist.cell(row=1, column=i, value=h)
            c.font = _font(XL_ACCENT, bold=True, size=9)
            c.fill = _fill(XL_SURFACE)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border_thin()
            ws_hist.column_dimensions[get_column_letter(i)].width = (
                22 if i == 1 else 12)
        ws_hist.row_dimensions[1].height = 18
    else:
        ws_hist = wb["History Log"]

    # Append history row
    hist_row = ws_hist.max_row + 1
    ws_hist.row_dimensions[hist_row].height = 15
    ws_hist.cell(row=hist_row, column=1, value=timestamp).font = _font(
        XL_MUTED, size=9)
    for col_idx, key in enumerate(indicators.keys(), 2):
        v = indicators[key].get("value")
        c = ws_hist.cell(row=hist_row, column=col_idx,
                         value=round(v, 4) if v is not None else None)
        c.font = _font(XL_TEXT, size=9)
        c.fill = _fill(XL_BG if hist_row % 2 == 0 else XL_SURFACE)
        c.border = _border_thin()
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Sheet 3: Indicator Reference ─────────────────────────────────────────
    if "Indicator Reference" not in wb.sheetnames:
        ws_ref = wb.create_sheet("Indicator Reference")
        ws_ref.sheet_properties.tabColor = XL_BEARISH
        ws_ref.sheet_view.showGridLines = False

        ref_headers = ["Key", "Label", "Measures", "Frequency",
                       "Source", "Bullish Threshold", "Bearish Threshold"]
        for i, h in enumerate(ref_headers, 1):
            c = ws_ref.cell(row=1, column=i, value=h)
            c.font = _font(XL_ACCENT, bold=True, size=9)
            c.fill = _fill(XL_SURFACE)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border_thin()
        col_ws = [14, 28, 38, 12, 10, 30, 30]
        for i, w in enumerate(col_ws, 1):
            ws_ref.column_dimensions[get_column_letter(i)].width = w
        ws_ref.row_dimensions[1].height = 18

        reference_data = [
            ("FED_FUNDS",      "Fed Funds Rate",            "FOMC target rate range",           "Per meeting", "FRED", "Rate cutting cycle", "Rate hiking cycle"),
            ("US_2Y",          "US 2Y Treasury",             "Short-end rate expectations",       "Daily",       "FRED", "Falling toward 3%",   "Rising above 5%"),
            ("US_10Y",         "US 10Y Treasury",            "Long-term growth/inflation",        "Daily",       "FRED", "Falling (flight to safety)", "Rising (tightening)"),
            ("US_30Y",         "US 30Y Treasury",            "Long-duration risk premium",        "Daily",       "FRED", "Falling",             "Rising above 5%"),
            ("SPREAD_2S10S",   "2s10s Yield Spread",         "Curve shape / recession signal",    "Daily",       "FRED", "> 0.5 (steepening)",  "< 0 (inverted)"),
            ("CPI_YOY",        "CPI YoY",                    "Headline inflation",                "Monthly",     "FRED", "Falling toward 2%",   "> 3.0% or rising"),
            ("CORE_CPI_YOY",   "Core CPI YoY",               "Inflation ex food & energy",        "Monthly",     "FRED", "< 2.5%",              "> 3.5%"),
            ("CORE_PCE_YOY",   "Core PCE YoY",               "Fed's preferred inflation gauge",   "Monthly",     "FRED", "< 2.0%",              "> 2.5%"),
            ("UNEMPLOYMENT",   "Unemployment Rate",          "Labor market slack",                "Monthly",     "FRED", "< 4.0%",              "Rising 0.2+pp"),
            ("INITIAL_CLAIMS", "Initial Jobless Claims",     "Weekly labor market pulse",         "Weekly",      "FRED", "< 220K",              "> 300K"),
            ("NFP",            "Nonfarm Payrolls MoM",       "Payroll job additions",             "Monthly",     "FRED", "> 150K",              "< 0"),
            ("ISM_MFG",        "ISM Manufacturing PMI",      "Factory sector activity",           "Monthly",     "FRED", "> 55",                "< 50"),
            ("CONSUMER_CONF",  "Consumer Confidence (Mich)", "Household sentiment",               "Monthly",     "FRED", "> 100",               "< 80 or -5pt drop"),
            ("IG_SPREAD",      "IG Credit Spread",           "Investment-grade risk premium",     "Daily",       "FRED", "< 100bps",            "> 200bps"),
            ("HY_SPREAD",      "HY Credit Spread",           "High-yield risk premium",           "Daily",       "FRED", "< 300bps",            "> 500bps"),
            ("VIX",            "VIX",                        "Implied equity volatility (fear)",  "Daily",       "yfinance", "< 15",            "> 20"),
            ("SP500",          "S&P 500",                    "US large-cap equity index",         "Daily",       "yfinance", "Above 50-day SMA", "Below 200-day SMA"),
            ("DOW",            "Dow Jones",                  "US blue-chip equity index",         "Daily",       "yfinance", "Uptrend",         "Downtrend"),
            ("NASDAQ",         "Nasdaq Composite",           "US tech-heavy equity index",        "Daily",       "yfinance", "Above 200-day SMA", "Below 200-day SMA"),
            ("FTSE100",        "FTSE 100",                   "UK large-cap equity index",         "Daily",       "yfinance", "Uptrend",         "Downtrend"),
            ("NIKKEI",         "Nikkei 225",                 "Japanese equity index",             "Daily",       "yfinance", "Uptrend",         "Downtrend"),
            ("OIL_WTI",        "WTI Crude Oil",              "US crude benchmark",                "Daily",       "yfinance", "$50-75 (goldilocks)", "> $90"),
            ("OIL_BRT",        "Brent Crude Oil",            "Global crude benchmark",            "Daily",       "yfinance", "$55-80",          "> $95"),
            ("GOLD",           "Gold",                       "Safe haven / inflation hedge",      "Daily",       "yfinance", "Rising steadily", "Sharp weekly drop > 3%"),
            ("COPPER",         "Copper",                     "Global growth proxy (Dr. Copper)",  "Daily",       "yfinance", "> $4/lb",         "< $3/lb"),
            ("DXY",            "DXY Dollar Index",           "USD vs basket of currencies",       "Daily",       "yfinance", "< 100 (weak $)",  "> 105 (strong $)"),
            ("PUT_CALL",       "Put/Call Ratio",             "Options market sentiment",          "Daily",       "Manual",  "< 0.7 or > 1.2 (contrarian)", "> 1.0"),
        ]
        for r_idx, row_data in enumerate(reference_data, 2):
            ws_ref.row_dimensions[r_idx].height = 16
            for c_idx, val in enumerate(row_data, 1):
                c = ws_ref.cell(row=r_idx, column=c_idx, value=val)
                c.font = _font(XL_TEXT if c_idx == 2 else XL_MUTED, size=9)
                c.fill = _fill(XL_SURFACE if r_idx % 2 == 0 else XL_BG)
                c.border = _border_thin()
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True)

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# 8. MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("\n⚡ Macro Pulse — Starting refresh...\n")

    # Load previous data
    previous_data = load_json(config.INDICATORS_FILE)
    rules = load_json(config.RULES_FILE)

    # Fetch data
    print("📡 Fetching FRED data...")
    fred_data = fetch_fred_data(previous_data)

    print("📡 Fetching yfinance data...")
    yf_data = fetch_yfinance_data(previous_data)

    # Merge
    indicators = {**fred_data, **yf_data}

    # Fallback: if a live fetch returned None but we have cached data, use it
    if previous_data:
        fallback_count = 0
        for key, ind in indicators.items():
            if ind.get("value") is None and isinstance(previous_data.get(key), dict):
                cached = previous_data[key]
                if cached.get("value") is not None:
                    ind["value"] = cached["value"]
                    ind["previous"] = cached.get("previous", ind.get("previous"))
                    ind["direction"] = cached.get("direction", "▬")
                    ind["history"] = cached.get("history", ind.get("history", []))
                    ind["value_date"] = cached.get("value_date")
                    ind["previous_date"] = cached.get("previous_date")
                    if "value_str" in cached:
                        ind["value_str"] = cached["value_str"]
                    fallback_count += 1
        if fallback_count:
            print(f"  ℹ  Used cached data for {fallback_count} indicator(s) where live fetch failed")

    # If first run, set previous = current (no direction)
    if not previous_data:
        for key, ind in indicators.items():
            if ind.get("previous") is None:
                ind["previous"] = ind.get("value")

    # Assign signals
    print("🔎 Assigning signals...")
    indicators = assign_signals(indicators, rules)

    # Generate commentary
    print("✍  Generating commentary...")
    for key, ind in indicators.items():
        ind["commentary"] = generate_commentary(key, ind)

    # Compute momentum
    print("📈 Computing momentum...")
    for key, ind in indicators.items():
        ind["momentum"] = compute_momentum(ind)

    # Compute 52-week range from history for indicators missing it
    for key, ind in indicators.items():
        history = ind.get("history", [])
        if history and len(history) > 0:
            values = [h[1] for h in history]
            if "high52w" not in ind:
                ind["high52w"] = max(values)
            if "low52w" not in ind:
                ind["low52w"] = min(values)

    # Timestamp
    now = datetime.now(timezone.utc)
    timestamp = now.strftime("%Y-%m-%d %H:%M UTC")

    # Generate summary panels
    print("📋 Generating summary panels...")
    daily_brief_text = generate_daily_brief(indicators)

    # Weekly wrap: only regenerate on Fri/Sat/Sun (weekday 4/5/6)
    if now.weekday() >= 4:
        weekly_wrap_text = generate_weekly_wrap(indicators, previous_data)
    elif previous_data.get("_weekly_wrap"):
        weekly_wrap_text = previous_data["_weekly_wrap"]
    else:
        weekly_wrap_text = "Weekly wrap generates on Friday. Check back then."

    forward_look_data = generate_forward_look(indicators)

    # Now add summaries to indicators for persistence
    indicators["_daily_brief"] = daily_brief_text
    indicators["_weekly_wrap"] = weekly_wrap_text
    indicators["_forward_look"] = forward_look_data

    # Save indicators
    indicators["_meta"] = {"last_updated": timestamp}
    save_json(config.INDICATORS_FILE, indicators)
    # Remove internal keys before passing to output generators
    meta = indicators.pop("_meta")
    daily_brief = indicators.pop("_daily_brief")
    weekly_wrap = indicators.pop("_weekly_wrap")
    forward_look = indicators.pop("_forward_look")

    # Generate HTML
    print("🌐 Generating HTML dashboard...")
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    html = generate_html(indicators, timestamp,
                         daily_brief=daily_brief,
                         weekly_wrap=weekly_wrap,
                         forward_look=forward_look)
    with open(config.HTML_OUTPUT, "w", encoding="utf-8") as f:
        f.write(html)

    # Generate Excel
    print("📊 Generating Excel workbook...")
    build_excel(indicators, timestamp)

    # Summary
    counts = {"BULLISH": 0, "BEARISH": 0, "NEUTRAL": 0}
    for ind in indicators.values():
        s = ind.get("signal", "NEUTRAL")
        counts[s] = counts.get(s, 0) + 1

    vix = indicators.get("VIX", {}).get("value")
    oil = indicators.get("OIL_WTI", {}).get("value")
    sp  = indicators.get("SP500", {}).get("value")
    vix_d = indicators.get("VIX", {}).get("direction", "▬")
    oil_d = indicators.get("OIL_WTI", {}).get("direction", "▬")
    sp_d  = indicators.get("SP500", {}).get("direction", "▬")

    print(f"""
╔══════════════════════════════════════════════════════════╗
║  ✓ Macro Pulse refreshed at {timestamp:<28}  ║
║                                                          ║
║  Bullish: {counts['BULLISH']:<3} │ Bearish: {counts['BEARISH']:<3} │ Neutral: {counts['NEUTRAL']:<3}            ║
║                                                          ║
║  VIX:   {(f"{vix:.2f}" if vix else "N/A"):<8} {vix_d}                                   ║
║  Oil:   ${(f"{oil:.2f}" if oil else "N/A"):<8} {oil_d}                                   ║
║  S&P:   {(f"{sp:,.0f}" if sp else "N/A"):<8} {sp_d}                                   ║
║                                                          ║
║  → {config.HTML_OUTPUT:<53}  ║
║  → {config.EXCEL_OUTPUT:<53}  ║
╚══════════════════════════════════════════════════════════╝
""")


if __name__ == "__main__":
    main()
